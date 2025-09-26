import csv, os, re
from time import time
# Import Helper Functions From Python Libraries
from openpyxl import load_workbook, Workbook

def convert_time_to_hours(time_str):
    """Convert a time string in 'HH:MM' format to hours as a float."""
    if not time_str:
        return -1
    try:
        hours, minutes = map(int, time_str.split(":"))
        return hours + minutes / 60
    except ValueError:
        return -1

def calculate_duration(start_time, end_time):
    """
    Calculate the duration between two time strings in hours.
    Returns -1 if the start_time or end_time is invalid.
    """
    start_hours = convert_time_to_hours(start_time)
    end_hours = convert_time_to_hours(end_time)
    if start_hours == -1 or end_hours == -1:
        return -1
    duration = end_hours - start_hours
    return duration

def calculate_event_staff_time(start_time, end_time, set_up_time, tear_down_time):
    """
    Calculate the total staff time for an event.
    Returns at minimum the event duration + any additional prep or post event time.
    If returns 0, the event is considered invalid.
    """
    # Event Duration) Calculate time between start_time and end_time
    event_duration = calculate_duration(start_time, end_time)
    # if event_duration is invalid, it means there was no listed event time
    if event_duration == -1:
        event_duration = 0

    # Prep Time) Calculate time between set_up_time and start time
    set_up_duration = calculate_duration(set_up_time, start_time)
    # if set_up_duration is invalid, it means there was no listed set up time
    if set_up_duration == -1:
        set_up_duration = 0

    # Post Time) Calculate time between end_time and tear_down_time
    tear_down_duration = calculate_duration(end_time, tear_down_time)
    # if tear_down_duration is invalid, it means there was no listed tear down time
    if tear_down_duration == -1:
        tear_down_duration = 0

    total_staff_time = event_duration + set_up_duration + tear_down_duration
    return total_staff_time

def create_dataset_template():
    # Create a Workbook
    wb = Workbook()

    # Give the Workbook The Following Worksheets
    workbook_setup = {  "EventInformation": [   "EventID", "Title", "Location", "Library Branch",
                                                "Event Organizer", "Presenter", "Audiences", 
                                                "Categories", "Internal Tags"],

                        "EventAudiences": ["EventID", "Audience"],
                        "EventCategories": ["EventID", "Category"],
                        "EventInternalTags": ["EventID", "Internal Tag"],

                        "EventTimes": [ "EventID", "Event Start Date", "Event End Date", "All Day Event",
                                        "Start Time", "End Time", "Set Up Time", "Tear Down Time", 
                                        "Duration", "Staff Time", "Raw Duration", "Different Dates"],

                        "EventParticipation": [ "EventID", "Registration Required", "In-Person Seats", 
                                                "Online Seats", "Confirmed Registrations", 
                                                "Waiting-List Registrations", "Cancelled Registrations", 
                                                "Anticipated Attendance", "Actual Attendance (In-Person)", 
                                                "Actual Attendance (Online)", "Confirmed Attendance"]
                     }

    # remove the first sheet
    wb.remove(wb["Sheet"])

    # Add Worksheets and Headers
    for sheet_name, headers in workbook_setup.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
    
    return wb

def find_data_source_file():
    # Search for the CSV file in the current directory
    # Pattern: starts with "lc_events_"
    csv_files = [f for f in os.listdir('.') if re.match(r'lc_events_.*\.csv$', f)]
    if not csv_files:
        print("No lc_events_*.csv file found in the current directory.")
        print("Could not update Matrix Map Dataset.")
        exit(1)
    else:
        print(f"Found data source file: {csv_files[0]}")
    return csv_files[0]

def read_csv_and_populate_workbook(wb, source_file):
    with open(source_file, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            # Skip header row
            if reader.line_num == 1:
                continue

            EventID = row[0]
            Title = row[1]
            # Description = row[2] *Not included to reduce unnecessary dataset size
            EventStartDate = row[3]
            EventEndDate = row[4] # *Not included to reduce unnecessary dataset size

            DifferentDates = (EventStartDate != EventEndDate)
            
            StartTime = row[5]      # Always Required
            EndTime = row[6]        # Sometimes Blank, if StartTime is "All Day Event"
            SetUpTime = row[7]      # Sometimes Blank
            TearDownTime = row[8]   # Sometimes Blank

            AllDayEvent = (StartTime == "All Day Event")

            Duration = calculate_duration(StartTime, EndTime) # Duration is -1 if no valid StartTime or EndTime
            StaffTime = calculate_event_staff_time(StartTime, EndTime, SetUpTime, TearDownTime)
            RawDuration = convert_time_to_hours(EndTime) - convert_time_to_hours(StartTime)

            Location = row[9]
            LibraryBranch = row[10]
            EventOrganizer = row[11]
            Presenter = row[12]
            Audiences = row[13]
            AudiencesList = [audience.strip() for audience in Audiences.split(",")]

            Categories = row[14]
            CategoriesList = [category.strip() for category in Categories.split(",")]

            # PublishingStatus = row[15] *Always set to "Published" Not included to reduce unnecessary dataset size
            InternalTags = row[16]
            InternalTagsList = [tag.strip() for tag in InternalTags.split(",")]
            # EventNote = row[17] *Not included to reduce unnecessary dataset size
            RegistrationRequired = row[18]
            InPersonSeats = row[19]
            OnlineSeats = row[20]
            ConfirmedRegistrations = row[21]
            WaitlistRegistrations = row[22]
            CancelledRegistrations = row[23]
            AnticipatedAttendance = row[24]
            ActualAttendance_InPerson = row[25]
            ActualAttendance_Online = row[26]
            ConfirmedAttendance = row[27]
            # EventURL = row[28] *Not included to reduce unnecessary dataset size

            # Add to Event Information Table
            ws = wb["EventInformation"]
            ws.append([EventID, Title, Location, LibraryBranch, EventOrganizer, Presenter, Audiences, Categories, InternalTags])

            # Add to Event Audiences Table
            ws = wb["EventAudiences"]
            for audience in AudiencesList:
                ws.append([EventID, audience])
            # if blank, put "Blank"
            if not AudiencesList:
                ws.append([EventID, "Blank"])

            # Add to Event Categories Table
            ws = wb["EventCategories"]
            for category in CategoriesList:
                ws.append([EventID, category])
            if not CategoriesList:
                ws.append([EventID, "Blank"])

            # Add to Event Internal Tags Table
            ws = wb["EventInternalTags"]
            for tag in InternalTagsList:
                ws.append([EventID, tag])
            if not InternalTagsList:
                ws.append([EventID, "Blank"])

            # Add to Event Times Table
            ws = wb["EventTimes"]
            ws.append([EventID, EventStartDate, EventEndDate, AllDayEvent, StartTime, EndTime, SetUpTime, TearDownTime, Duration, StaffTime, RawDuration, DifferentDates])

            # Add to Event Participation Table
            ws = wb["EventParticipation"]
            ws.append([EventID, RegistrationRequired, InPersonSeats, OnlineSeats, ConfirmedRegistrations, 
                       WaitlistRegistrations, CancelledRegistrations, AnticipatedAttendance, ActualAttendance_InPerson, 
                       ActualAttendance_Online, ConfirmedAttendance])

    # close file
    f.close()

def main():
    # Step 1: Search for Data Source File 
    source_file = find_data_source_file()
    print(f"Using data source file: {source_file}")

    # Step 2: Create Dataset Template
    wb = create_dataset_template()
    print(f"Created dataset template: {wb}")

    # Step 3: Check for Existing Dataset and delete it
    if os.path.exists("MatrixMapDataset.xlsx"):
        os.remove("MatrixMapDataset.xlsx")
        print("Deleted existing dataset file.")

    # Step 4: Read CSV and Populate Workbook
    read_csv_and_populate_workbook(wb, source_file)
    print("Populated workbook with data from CSV.")

    # Step 5: Save the Workbook
    wb.save("MatrixMapDataset.xlsx")
    print("Saved workbook as 'MatrixMapDataset.xlsx'.")

    # Step 6: Close the workbook
    wb.close()

import time
if __name__ == "__main__":
    '''
    This program organizes the LibCal Events Exported Dataset into a structured Excel workbook 
    which facilitates data analysis and visualization when imported into Power BI.

    Data Entry Validation Tool:
    The Power BI tool I'm making will visualize potential data entry errors and inconsistencies.
    1) Display information of EventIDs and their Titles where start and end dates differ
    2) Display information of EventIDs and their Titles with a slicer for Raw Duration
    3) Display information of EventIDs and their Titles where the Library Branch or Location is Inaccurate

    Title Groupings:
    Since event titles may not always be consistent, I need to do some text analysis:

    Solution 1) Add new worksheet titles EventGroupings
    Headers = ["Group Title", "Title"]
    Examples of rows:
    ["Storytime", "Storytime"]
    ["Storytime", "Story Time"]
    ["Storytime", "Stories for Kids"]
    * Pros: Allows data to be filtered by group simply in Power BI
    * Cons: Unless a pattern can be identified with an algorithm, manual implementation will be required!

    Solution 2) Use Machine Learning for Title Classification
    - Train a model to classify event titles into predefined groups
    - Use techniques like Natural Language Processing (NLP) to analyze title text
    - Automate the grouping process, reducing manual effort

    Solution 3) Implement a Hybrid Approach
    - Combine rule-based and machine learning techniques for better accuracy
    - Use initial rules to filter and group titles, then apply ML for final classification

    Solution 4) Manual Review and Adjustment
    - After automated grouping, conduct a manual review to catch any misclassifications
    - Allow for user feedback to improve the grouping process over time

    '''
    start_time = time.time()
    main()
    end_time = time.time()
    print(f"Finished in {end_time - start_time:.3f} seconds.")