import csv
# Import Helper Functions From Python Libraries
from openpyxl import load_workbook, Workbook

# Create a Workbook
wb = Workbook()

# Give the Workbook The Following Worksheets
workbook_setup = {"EventInformation": ["EventID", "Title", "Event Date", "Duration", "Staff Time", "Location", "Library Branch", "Event Organizer", "Presenter", "Audiences", "Categories", "Publishing Status", "Internal Tags", "Registration Required", "In-Person Seats", "Online Seats", "Confirmed Registrations", "Waiting-List Registrations", "Cancelled Registrations", "Anticipated Attendance", "Actual Attendance (In-Person)", "Actual Attendance (Online)", "Confirmed Attendance"],
                  "EventAudiences": ["EventID", "Audience"],
                  "EventCategories": ["EventID", "Category"],
                  "EventInternalTags": ["EventID", "Internal Tag"]}

# remove the first sheet
wb.remove(wb["Sheet"])

for sheet_name, headers in workbook_setup.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.append(headers)

with open("Programs2024.csv", newline="", encoding="utf-8") as f:
    reader = csv.reader(f)
    for row in reader:
        # Check if already Published
        Published = row[15]
        print(Published)
        if Published != "Published":
            continue

        EventID = row[0]
        Title = row[1]
        EventDate = row[3]
        StartTime = row[5]
        EndTime = row[6]
        Duration = "Not Yet Calculated" # Placeholder for duration calculation
        SetUpTime = row[7]
        TearDownTime = row[8]
        StaffTime = "Not Yet Calculated" # Placeholder for staff time calculation
        Location = row[9]
        LibraryBranch = row[10]
        EventOrganizer = row[11]
        Presenter = row[12]

        Audiences = row[13]
        AudiencesList = [audience.strip() for audience in Audiences.split(",")]

        Categories = row[14]
        CategoriesList = [category.strip() for category in Categories.split(",")]

        InternalTags = row[16]
        InternalTagsList = [tag.strip() for tag in InternalTags.split(",")]

        RegistrationRequired = row[18]
        InPersonSeats = row[19]
        OnlineSeats = row[20]
        ConfirmedRegistrations = row[21]
        WaitlistRegistrations = row[22]
        CancelledRegistrations = row[23]
        AnticipatedAttendance = row[24]
        ActualAttendance_InPerson = row[25]
        ActualAttendance_Online = row[26]
        ConfirmedAttendance = row[27]

        # Add to Event Information Table
        ws = wb["EventInformation"]
        ws.append([EventID, Title, EventDate, Duration, StaffTime, Location, LibraryBranch, EventOrganizer, Presenter, Audiences, Categories, Published, InternalTags, RegistrationRequired, InPersonSeats, OnlineSeats, ConfirmedRegistrations, WaitlistRegistrations, CancelledRegistrations, AnticipatedAttendance, ActualAttendance_InPerson, ActualAttendance_Online, ConfirmedAttendance])

        # Add to Event Audiences Table
        ws = wb["EventAudiences"]
        for audience in AudiencesList:
            ws.append([EventID, audience])

        # Add to Event Categories Table
        ws = wb["EventCategories"]
        for category in CategoriesList:
            ws.append([EventID, category])

        # Add to Event Internal Tags Table
        ws = wb["EventInternalTags"]
        for tag in InternalTagsList:
            ws.append([EventID, tag])
            print(f"Added {EventID} with tag {tag}")

    # Save the workbook
    wb.save("EventsDatasetForPowerBI.xlsx")
