import csv
from openpyxl import load_workbook, Workbook
import os
import sys
from datetime import datetime

def safe_split(value, delimiter=','):
    """Safely split a string, handling empty values"""
    if not value or str(value).strip() == '':
        return []
    return [item.strip() for item in str(value).split(delimiter) if item.strip()]

def safe_get(row, index, default=''):
    """Safely get a value from a list with bounds checking"""
    try:
        return row[index] if index < len(row) else default
    except (IndexError, TypeError):
        return default

def calculate_duration(start_time, end_time):
    """Calculate duration between start and end time"""
    try:
        if not start_time or not end_time:
            return "Not Available"
        
        # Simple time calculation - assumes format like "10:00 AM"
        # You may need to adjust this based on your actual time format
        start = datetime.strptime(start_time, "%I:%M %p")
        end = datetime.strptime(end_time, "%I:%M %p")
        
        duration = end - start
        hours = duration.seconds // 3600
        minutes = (duration.seconds % 3600) // 60
        
        if hours > 0:
            return f"{hours}h {minutes}m"
        else:
            return f"{minutes}m"
    except (ValueError, AttributeError):
        return "Unable to Calculate"

def calculate_staff_time(setup_time, teardown_time, duration_str):
    """Calculate total staff time including setup and teardown"""
    try:
        setup = float(setup_time) if setup_time and setup_time != '' else 0
        teardown = float(teardown_time) if teardown_time and teardown_time != '' else 0
        
        # Extract duration in minutes (simplified)
        duration_minutes = 0
        if 'h' in duration_str and 'm' in duration_str:
            parts = duration_str.replace('h', '').replace('m', '').split()
            if len(parts) == 2:
                duration_minutes = int(parts[0]) * 60 + int(parts[1])
        elif 'm' in duration_str:
            duration_minutes = int(duration_str.replace('m', ''))
        
        total_minutes = setup + teardown + duration_minutes
        hours = int(total_minutes // 60)
        minutes = int(total_minutes % 60)
        
        if hours > 0:
            return f"{hours}h {minutes}m"
        else:
            return f"{minutes}m"
    except (ValueError, AttributeError):
        return "Unable to Calculate"

def main():
    # Define expected CSV column mapping
    CSV_COLUMNS = {
        'EventID': 0,
        'Title': 1,
        'EventDate': 4,
        'StartTime': 5,
        'EndTime': 6,
        'SetUpTime': 8,
        'TearDownTime': 9,
        'Location': 10,
        'LibraryBranch': 11,
        'EventOrganizer': 12,
        'Presenter': 13,
        'Audiences': 14,
        'Categories': 15,
        'Published': 16,
        'InternalTags': 17,
        'RegistrationRequired': 19,
        'InPersonSeats': 20,
        'OnlineSeats': 21,
        'ConfirmedRegistrations': 22,
        'WaitlistRegistrations': 23,
        'CancelledRegistrations': 24,
        'AnticipatedAttendance': 25,
        'ActualAttendance_InPerson': 26,
        'ActualAttendance_Online': 27,
        'ConfirmedAttendance': 28
    }

    # Create a Workbook
    wb = Workbook()

    # Define worksheet structure
    workbook_setup = {
        "EventInformation": [
            "EventID", "Title", "Event Date", "Duration", "Staff Time", "Location", 
            "Library Branch", "Event Organizer", "Presenter", "Publishing Status", 
            "Registration Required", "In-Person Seats", "Online Seats", 
            "Confirmed Registrations", "Waiting-List Registrations", 
            "Cancelled Registrations", "Anticipated Attendance", 
            "Actual Attendance (In-Person)", "Actual Attendance (Online)", 
            "Confirmed Attendance"
        ],
        "EventAudiences": ["EventID", "Audience"],
        "EventCategories": ["EventID", "Category"],
        "EventInternalTags": ["EventID", "Internal Tag"]
    }

    # Remove the default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create worksheets with headers
    for sheet_name, headers in workbook_setup.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

    # Check if CSV file exists
    csv_filename = "Programs2024.csv"
    if not os.path.exists(csv_filename):
        print(f"Error: {csv_filename} not found in current directory")
        print(f"Current directory: {os.getcwd()}")
        return False

    try:
        processed_count = 0
        skipped_count = 0
        
        with open(csv_filename, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            
            # Skip header row if it exists
            try:
                first_row = next(reader)
                # Check if first row looks like headers
                if not first_row[0].isdigit():  # Assuming EventID is numeric
                    print("Detected header row, skipping...")
                else:
                    # Process the first row if it's not headers
                    process_row(wb, first_row, CSV_COLUMNS)
                    processed_count += 1
            except StopIteration:
                print("CSV file appears to be empty")
                return False

            for row_num, row in enumerate(reader, start=2):  # Start at 2 since we may have skipped header
                try:
                    if len(row) < max(CSV_COLUMNS.values()) + 1:
                        print(f"Warning: Row {row_num} has insufficient columns ({len(row)}), skipping")
                        skipped_count += 1
                        continue

                    # Check if event is published
                    published = safe_get(row, CSV_COLUMNS['Published'])
                    if published != "Published":
                        print(f"Skipping unpublished event in row {row_num}: {safe_get(row, CSV_COLUMNS['Title'])}")
                        skipped_count += 1
                        continue

                    process_row(wb, row, CSV_COLUMNS)
                    processed_count += 1

                except Exception as e:
                    print(f"Error processing row {row_num}: {e}")
                    skipped_count += 1
                    continue

        print(f"\nProcessing complete!")
        print(f"Processed: {processed_count} events")
        print(f"Skipped: {skipped_count} events")

    except FileNotFoundError:
        print(f"Error: Could not find {csv_filename}")
        return False
    except PermissionError:
        print(f"Error: Permission denied reading {csv_filename}")
        return False
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return False

    # Save the workbook
    try:
        output_filename = "ProcessedEvents.xlsx"
        wb.save(output_filename)
        print(f"Workbook saved as: {output_filename}")
        return True
    except PermissionError:
        print(f"Error: Could not save {output_filename}. File may be open in another program.")
        return False
    except Exception as e:
        print(f"Error saving workbook: {e}")
        return False

def process_row(wb, row, columns):
    """Process a single row of CSV data"""
    
    # Extract basic event information
    event_id = safe_get(row, columns['EventID'])
    title = safe_get(row, columns['Title'])
    event_date = safe_get(row, columns['EventDate'])
    start_time = safe_get(row, columns['StartTime'])
    end_time = safe_get(row, columns['EndTime'])
    setup_time = safe_get(row, columns['SetUpTime'])
    teardown_time = safe_get(row, columns['TearDownTime'])
    location = safe_get(row, columns['Location'])
    library_branch = safe_get(row, columns['LibraryBranch'])
    event_organizer = safe_get(row, columns['EventOrganizer'])
    presenter = safe_get(row, columns['Presenter'])
    published = safe_get(row, columns['Published'])

    # Calculate duration and staff time
    duration = calculate_duration(start_time, end_time)
    staff_time = calculate_staff_time(setup_time, teardown_time, duration)

    # Extract list fields
    audiences = safe_split(safe_get(row, columns['Audiences']))
    categories = safe_split(safe_get(row, columns['Categories']))
    internal_tags = safe_split(safe_get(row, columns['InternalTags']))

    # Extract registration and attendance data
    registration_required = safe_get(row, columns['RegistrationRequired'])
    in_person_seats = safe_get(row, columns['InPersonSeats'])
    online_seats = safe_get(row, columns['OnlineSeats'])
    confirmed_registrations = safe_get(row, columns['ConfirmedRegistrations'])
    waitlist_registrations = safe_get(row, columns['WaitlistRegistrations'])
    cancelled_registrations = safe_get(row, columns['CancelledRegistrations'])
    anticipated_attendance = safe_get(row, columns['AnticipatedAttendance'])
    actual_attendance_in_person = safe_get(row, columns['ActualAttendance_InPerson'])
    actual_attendance_online = safe_get(row, columns['ActualAttendance_Online'])
    confirmed_attendance = safe_get(row, columns['ConfirmedAttendance'])

    # Add to EventInformation worksheet
    ws_info = wb["EventInformation"]
    ws_info.append([
        event_id, title, event_date, duration, staff_time, location, library_branch,
        event_organizer, presenter, published, registration_required, in_person_seats,
        online_seats, confirmed_registrations, waitlist_registrations, 
        cancelled_registrations, anticipated_attendance, actual_attendance_in_person,
        actual_attendance_online, confirmed_attendance
    ])

    # Add to EventAudiences worksheet
    ws_audiences = wb["EventAudiences"]
    for audience in audiences:
        ws_audiences.append([event_id, audience])

    # Add to EventCategories worksheet
    ws_categories = wb["EventCategories"]
    for category in categories:
        ws_categories.append([event_id, category])

    # Add to EventInternalTags worksheet
    ws_tags = wb["EventInternalTags"]
    for tag in internal_tags:
        ws_tags.append([event_id, tag])

if __name__ == "__main__":
    success = main()
    if not success:
        sys.exit(1)