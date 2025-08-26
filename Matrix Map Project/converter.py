import csv
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_excel_template():
    """
    Create Excel template with all worksheets and headers
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define worksheet structure with headers
    worksheet_templates = {
        "Summary": ["Metric", "Registration_Type", "Total", "In_Person", "Online"],
        "Monthly_Distribution": ["Month_Year", "Metric", "Value"],
        "Day_of_Week_Distribution": ["Day_of_Week", "Metric", "Value"],
        "Hour_Distribution": ["Hour", "Metric", "Value"],
        "Daily_Hourly_Distribution": ["Day_of_Week", "Hour", "Value"],
        "Library_Branch_Distribution": ["Month_Year", "Library_Branch", "Events"],
        "Audience_Distribution": ["Month_Year", "Audience_Type", "Events"],
        "Category_Distribution": ["Month_Year", "Category", "Events"],
        "Manatee_Library_Events": ["Month_Year", "Metric", "Value"],
        "Staff_Training_Events": ["Month_Year", "Metric", "Value"],
        "Other_Calendar_Events": ["Month_Year", "Metric", "Value"]
    }
    
    # Create worksheets with headers
    for sheet_name, headers in worksheet_templates.items():
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            # Format headers
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(len(header) + 2, 12)
    
    return wb

def parse_events_calendar_csv(csv_file_path, output_excel_path):
    """
    Convert EventsCalendar2024.csv to Power BI optimized Excel workbook
    """
    
    # Create Excel template
    print("Creating Excel template...")
    wb = create_excel_template()
    
    # Read the CSV file
    print("Reading CSV file...")
    with open(csv_file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Split content into sections
    sections = content.split('\n\n')
    
    # Process each section
    print("Processing data sections...")
    for section in sections:
        if not section.strip():
            continue
            
        lines = section.strip().split('\n')
        if not lines:
            continue
            
        # Get section header
        header = lines[0].strip()
        
        # Skip if no data rows
        if len(lines) < 2:
            continue
            
        # Parse different section types
        if header == "Summary":
            populate_summary_sheet(wb, lines[1:])
        elif header == "Monthly Distribution":
            populate_monthly_distribution_sheet(wb, lines[1:])
        elif header == "Day of the Week Distribution":
            populate_day_of_week_sheet(wb, lines[1:])
        elif header == "Hour of the Day Distribution":
            populate_hour_distribution_sheet(wb, lines[1:])
        elif header == "Daily/Hourly Distribution":
            populate_daily_hourly_sheet(wb, lines[1:])
        elif "Library Branch Distribution" in header:
            populate_library_branch_sheet(wb, lines[1:])
        elif header == "Audience Distribution":
            populate_audience_sheet(wb, lines[1:])
        elif header == "Category Distribution":
            populate_category_sheet(wb, lines[1:])
        elif "Individual Calendars Monthly Breakdown" in header:
            process_individual_calendars(wb, section)
    
    # Final formatting and optimization
    print("Applying final formatting...")
    optimize_workbook(wb)
    
    # Save the workbook
    wb.save(output_excel_path)
    print(f"Excel file saved to: {output_excel_path}")
    return wb

def populate_summary_sheet(wb, data_lines):
    """Populate Summary worksheet"""
    ws = wb["Summary"]
    row_num = 2
    
    for line in data_lines:
        if line.strip():
            parts = [part.strip() for part in line.split(',')]
            if len(parts) >= 4:
                for col_num, value in enumerate(parts, 1):
                    # Convert numeric values
                    if col_num > 2:  # Total, In_Person, Online columns
                        try:
                            numeric_value = int(value) if value.isdigit() else (value if value == '-' else 0)
                            ws.cell(row=row_num, column=col_num, value=numeric_value)
                        except:
                            ws.cell(row=row_num, column=col_num, value=value)
                    else:
                        ws.cell(row=row_num, column=col_num, value=value)
                row_num += 1

def populate_monthly_distribution_sheet(wb, data_lines):
    """Populate Monthly Distribution worksheet"""
    ws = wb["Monthly_Distribution"]
    row_num = 2
    
    # Parse header row to get months
    header_line = data_lines[0]
    months = [month.strip() for month in header_line.split(',')[1:]]
    
    # Process each metric row
    for line in data_lines[1:]:
        if line.strip():
            parts = [part.strip() for part in line.split(',')]
            metric = parts[0]
            values = parts[1:]
            
            # Create normalized rows
            for i, value in enumerate(values):
                if i < len(months) and value:
                    try:
                        numeric_value = int(value) if value.isdigit() else 0
                        ws.cell(row=row_num, column=1, value=months[i])
                        ws.cell(row=row_num, column=2, value=metric)
                        ws.cell(row=row_num, column=3, value=numeric_value)
                        row_num += 1
                    except:
                        ws.cell(row=row_num, column=1, value=months[i])
                        ws.cell(row=row_num, column=2, value=metric)
                        ws.cell(row=row_num, column=3, value=value)
                        row_num += 1

def populate_day_of_week_sheet(wb, data_lines):
    """Populate Day of Week Distribution worksheet"""
    ws = wb["Day_of_Week_Distribution"]
    row_num = 2
    
    # Get days from header
    header_line = data_lines[0]
    days = [day.strip() for day in header_line.split(',')[1:]]
    
    # Process each metric
    for line in data_lines[1:]:
        if line.strip():
            parts = [part.strip() for part in line.split(',')]
            metric = parts[0]
            values = parts[1:]
            
            for i, value in enumerate(values):
                if i < len(days) and value:
                    try:
                        numeric_value = int(value) if value.isdigit() else 0
                        ws.cell(row=row_num, column=1, value=days[i])
                        ws.cell(row=row_num, column=2, value=metric)
                        ws.cell(row=row_num, column=3, value=numeric_value)
                        row_num += 1
                    except:
                        continue

def populate_hour_distribution_sheet(wb, data_lines):
    """Populate Hour Distribution worksheet"""
    ws = wb["Hour_Distribution"]
    row_num = 2
    
    # Get hours from header
    header_line = data_lines[0]
    hours = [hour.strip() for hour in header_line.split(',')[1:]]
    
    # Process each metric
    for line in data_lines[1:]:
        if line.strip():
            parts = [part.strip() for part in line.split(',')]
            metric = parts[0]
            values = parts[1:]
            
            for i, value in enumerate(values):
                if i < len(hours) and value:
                    try:
                        numeric_value = int(value) if value.isdigit() else 0
                        ws.cell(row=row_num, column=1, value=int(hours[i]))
                        ws.cell(row=row_num, column=2, value=metric)
                        ws.cell(row=row_num, column=3, value=numeric_value)
                        row_num += 1
                    except:
                        continue

def populate_daily_hourly_sheet(wb, data_lines):
    """Populate Daily/Hourly Distribution worksheet"""
    ws = wb["Daily_Hourly_Distribution"]
    row_num = 2
    
    # Get hours from header
    header_line = data_lines[0]
    hours = [hour.strip() for hour in header_line.split(',')[1:]]
    
    # Process each day
    for line in data_lines[1:]:
        if line.strip():
            parts = [part.strip() for part in line.split(',')]
            day = parts[0]
            values = parts[1:]
            
            for i, value in enumerate(values):
                if i < len(hours) and value:
                    try:
                        numeric_value = int(value) if value.isdigit() else 0
                        if numeric_value > 0:  # Only include non-zero values
                            ws.cell(row=row_num, column=1, value=day)
                            ws.cell(row=row_num, column=2, value=int(hours[i]))
                            ws.cell(row=row_num, column=3, value=numeric_value)
                            row_num += 1
                    except:
                        continue

def populate_library_branch_sheet(wb, data_lines):
    """Populate Library Branch Distribution worksheet"""
    ws = wb["Library_Branch_Distribution"]
    row_num = 2
    
    # Get months from header
    header_line = data_lines[0]
    months = [month.strip().strip('"') for month in header_line.split(',')[1:]]
    
    # Process each branch
    for line in data_lines[1:]:
        if line.strip():
            parts = [part.strip().strip('"') for part in line.split(',')]
            branch = parts[0]
            values = parts[1:]
            
            for i, value in enumerate(values):
                if i < len(months) and value:
                    try:
                        numeric_value = int(value) if value.isdigit() else 0
                        ws.cell(row=row_num, column=1, value=months[i])
                        ws.cell(row=row_num, column=2, value=branch)
                        ws.cell(row=row_num, column=3, value=numeric_value)
                        row_num += 1
                    except:
                        continue

def populate_audience_sheet(wb, data_lines):
    """Populate Audience Distribution worksheet"""
    ws = wb["Audience_Distribution"]
    row_num = 2
    
    # Get months from header
    header_line = data_lines[0]
    months = [month.strip().strip('"') for month in header_line.split(',')[1:]]
    
    # Process each audience type
    for line in data_lines[1:]:
        if line.strip():
            parts = [part.strip().strip('"') for part in line.split(',')]
            audience = parts[0]
            values = parts[1:]
            
            for i, value in enumerate(values):
                if i < len(months) and value:
                    try:
                        numeric_value = int(value) if value.isdigit() else 0
                        ws.cell(row=row_num, column=1, value=months[i])
                        ws.cell(row=row_num, column=2, value=audience)
                        ws.cell(row=row_num, column=3, value=numeric_value)
                        row_num += 1
                    except:
                        continue

def populate_category_sheet(wb, data_lines):
    """Populate Category Distribution worksheet"""
    ws = wb["Category_Distribution"]
    row_num = 2
    
    # Get months from header
    header_line = data_lines[0]
    months = [month.strip().strip('"') for month in header_line.split(',')[1:]]
    
    # Process each category
    for line in data_lines[1:]:
        if line.strip():
            parts = [part.strip().strip('"') for part in line.split(',')]
            category = parts[0]
            values = parts[1:]
            
            for i, value in enumerate(values):
                if i < len(months) and value:
                    try:
                        numeric_value = int(value) if value.isdigit() else 0
                        ws.cell(row=row_num, column=1, value=months[i])
                        ws.cell(row=row_num, column=2, value=category)
                        ws.cell(row=row_num, column=3, value=numeric_value)
                        row_num += 1
                    except:
                        continue

def process_individual_calendars(wb, section_content):
    """Process Individual Calendars section"""
    lines = section_content.strip().split('\n')
    
    current_sheet = None
    months = []
    header_found = False
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Detect which calendar type based on content
        if 'Manatee Library Events Monthly Distribution' in line:
            current_sheet = wb["Manatee_Library_Events"]
            header_found = False
        elif 'Staff Training  Monthly Distribution' in line:
            current_sheet = wb["Staff_Training_Events"]
            header_found = False
        elif 'Monthly Distribution' in line and current_sheet is None:
            current_sheet = wb["Other_Calendar_Events"]
            header_found = False
        
        # Check if this is the month header row
        if line.startswith('Month/Year') and current_sheet is not None:
            months = [month.strip() for month in line.split(',')[1:]]
            header_found = True
            continue
        
        # Process data rows
        if header_found and current_sheet is not None and ',' in line and not line.startswith('Month/Year'):
            parts = [part.strip() for part in line.split(',')]
            metric = parts[0]
            values = parts[1:]
            
            # Find the next empty row
            row_num = current_sheet.max_row + 1
            
            for i, value in enumerate(values):
                if i < len(months) and value:
                    try:
                        numeric_value = int(value) if value.isdigit() else 0
                        current_sheet.cell(row=row_num, column=1, value=months[i])
                        current_sheet.cell(row=row_num, column=2, value=metric)
                        current_sheet.cell(row=row_num, column=3, value=numeric_value)
                        row_num += 1
                    except:
                        continue

def optimize_workbook(wb):
    """Apply final formatting and optimization to workbook"""
    for ws in wb.worksheets:
        # Apply alternating row colors for better readability
        light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_num in range(2, ws.max_row + 1):
            if row_num % 2 == 0:
                for col_num in range(1, ws.max_column + 1):
                    ws.cell(row=row_num, column=col_num).fill = light_fill
        
        # Center align all data
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-fit column widths based on content
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

def main():
    """Main function to run the conversion"""
    # Define file paths
    input_csv = "EventsCalendar2024.csv"
    output_excel = "EventsCalendar2024_PowerBI.xlsx"
    
    # Check if input file exists
    if not os.path.exists(input_csv):
        print(f"Error: Input file '{input_csv}' not found.")
        print("Please ensure the CSV file is in the same directory as this script.")
        return
    
    try:
        # Convert the CSV to Excel
        wb = parse_events_calendar_csv(input_csv, output_excel)
        
        # Print summary
        if os.path.exists(output_excel):
            print(f"\n Conversion completed successfully!")
            print(f" Created {len(wb.sheetnames)} worksheets:")
            
            for ws in wb.worksheets:
                row_count = ws.max_row - 1  # Subtract header row
                print(f"    {ws.title}: {row_count} data rows")
            
            wb.close()
            
            print(f"\n The file '{output_excel}' is now ready for Power BI import!")
            print(" Each worksheet represents a different data dimension for analysis.")
            print(" Import each sheet as a separate table in Power BI for comprehensive dashboards.")
        
    except Exception as e:
        print(f"X Error during conversion: {str(e)}")
        print("Please check your CSV file format and try again.")

if __name__ == "__main__":
    main()