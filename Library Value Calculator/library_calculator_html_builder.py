from openpyxl import load_workbook
import os

def generate_templates():
    # Step 1) Open for CalculatorSettings.xlsx
    script_directory = os.getcwd()
    wb = None
    for filename in os.listdir(script_directory):
        if filename.startswith("~$"):
            continue
        if filename == "Calculator Settings.xlsx":
            full_path = os.path.join(script_directory, filename)
            
            # Add debugging info
            print(f"Found file: {full_path}")
            print(f"File exists: {os.path.exists(full_path)}")
            print(f"File size: {os.path.getsize(full_path)} bytes")
            
            # Check if file is actually an Excel file
            try:
                wb = load_workbook(full_path, data_only=True)
                print("Successfully opened the file!")
                break
            except Exception as e:
                print(f"Error opening file: {e}")
                # Try without data_only flag
                try:
                    wb = load_workbook(full_path)
                    print("Opened without data_only=True")
                    return
                except Exception as e2:
                    print(f"Still failed without data_only=True: {e2}")
                    return
    
    # Step 2) Get Information for Each Service
    Calculator_Settings_Dictionary = dict()
    row = 2
    while wb["Values Per Service"]["A" + str(row)].value is not None:
        Service = wb["Values Per Service"]["A"+str(row)].value
        Value_Per_Service = wb["Values Per Service"]["B"+str(row)].value
        Value_Explained = wb["Values Per Service"]["C"+str(row)].value
        Calculator_Settings_Dictionary[Service] = [Value_Per_Service, Value_Explained]
        row += 1

    # Step 3) Get Fiscal Year Data for Each Service
    Fiscal_Year_Dictionary = dict()
    def get_column_letter(col_num):
        letter = ''
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    column = 2
    column_letter = get_column_letter(column)

    while wb["Fiscal Year Values"][column_letter + "2"].value is not None:
        Fiscal_Year = wb["Fiscal Year Values"][column_letter + "2"].value
        
        row = 3
        while wb["Fiscal Year Values"]["A" + str(row)].value is not None:
            Service = wb["Fiscal Year Values"]["A" + str(row)].value
            if Service in Calculator_Settings_Dictionary.keys():
                Total = wb["Fiscal Year Values"][column_letter + str(row)].value

                if Fiscal_Year not in Fiscal_Year_Dictionary:
                    Fiscal_Year_Dictionary[Fiscal_Year] = [(Service, Total)]
                else:    
                    Fiscal_Year_Dictionary[Fiscal_Year].append((Service, Total))
            else:
                print(f"{Service} was not found as a Service on the 'Values Per Service' Worksheet")
            row += 1

        column += 1
        column_letter = get_column_letter(column)

    # Step 4) Recreate the Script

    # Step 4.1) productValues
    CODE_productValues = "\tconst productValues = {\n"
    for key, value in Calculator_Settings_Dictionary.items():
        CODE_productValues += f"\t\t\"{key}\":{value[0]},\n"
    CODE_productValues += "\t};\n\n"

    # Step 4.2) fiscalYearData
    CODE_fiscalYearData = "\tconst fiscalYearData = {\n"
    for key, value in Fiscal_Year_Dictionary.items():
        CODE_fiscalYearData += f"\t\t\"{key}\": " + "{\n"
        for service_info in value:
            Service = service_info[0]
            Total = service_info[1]
            CODE_fiscalYearData += f"\t\t\t\"{Service}\":{Total},\n"
        CODE_fiscalYearData += "\t\t},\n"
    CODE_fiscalYearData += "\t};\n"

    def get_style():
            return '''
<style>
    .container {
        color: #333333;
        text-align: left;
        max-width: 1400px;
        margin: 0 auto;
        padding: 20px;
    }

    table {
        width: 100%;
        border-collapse: collapse;
        border: 2px solid #415364;
        margin-bottom: 30px;
    }

    th {
        background-color: #415364;
        color: white;
        font-weight: bold;
        padding: 12px 20px;
        border: 1px solid #415364;
    }

    th:nth-child(4),
    th:nth-child(5) {
        text-align: right;
    }

    td {
        padding: 12px 20px;
        border: 1px solid #415364;
    }

    tbody tr:nth-child(even) {
        background-color: #e5e9ed;
    }

    tfoot th {
        font-weight: bold;
        border: 1px solid #415364;
        background-color: #415364;
        color: white;
    }

    input[type="number"] {
        width: 120px;
        padding: 8px;
        border: 1px solid #ccc;
        border-radius: 4px;
        font-size: 14px;
    }

    .value-cell {
        padding: 12px 20px;
        text-align: right;
        font-weight: bold;
    }

    .values-explained {
        text-align: left;
        /* font-weight: bold; */
        /* color: #666; */
        /* font-size: 14px; */
    }

    .values-explained-column {
        display: none;
    }

    .values-explained-column.visible {
        display: table-cell;
    }

    .value-per-use-column {
        display: none;
        text-align: right;
        font-weight: bold;
        padding: 12px 20px;
        /* border: 1px solid #415364; */
    }

    .value-per-use-column.visible {
        display: table-cell;
    }

    #total-value {
        text-align: right;
        font-size: 18px;
    }

    .buttons-container {
        margin-bottom: 30px;
        text-align: center;
    }

    .buttons-container h2 {
        color: #415364;
        margin-bottom: 15px;
    }

    .fy-buttons-row {
        margin-bottom: 15px;
    }

    .control-buttons-row {
        margin-top: 15px;
    }

    .fy-button {
        background-color: #415364;
        color: white;
        border: none;
        padding: 12px 24px;
        margin: 0 10px 10px 10px;
        border-radius: 6px;
        font-size: 16px;
        font-weight: bold;
        cursor: pointer;
        transition: background-color 0.3s ease;
    }

    .fy-button:hover {
        background-color: #2c3742;
    }

    .fy-button:active {
        background-color: #2c3742;
    }

    .clear-button, .toggle-explanation-button {
        background-color: #d15e14;
        color: white;
        border: none;
        padding: 10px 20px;
        margin: 0 10px 10px 10px;
        border-radius: 6px;
        font-size: 14px;
        cursor: pointer;
        transition: background-color 0.3s ease;
    }

    .clear-button:hover, .toggle-explanation-button:hover {
        background-color: #9B2E21;
    }

    .toggle-explanation-button {
        background-color: #415364;
    }

    .toggle-explanation-button:hover {
        background-color: #2c3742;
    }

    /* Responsive adjustments for smaller screens */
    @media (max-width: 768px) {
        .container {
            padding: 10px;
        }
        
        input[type="number"] {
            width: 80px;
        }
        
        th, td {
            padding: 8px 10px;
            font-size: 13px;
        }
        
        .values-explained {
            font-size: 12px;
        }
    }
</style>
'''
    def get_script():
        CODE_script = "<script>\n"
        CODE_script += CODE_productValues
        CODE_script += CODE_fiscalYearData
        CODE_script += '''
    let inputs;
    let totalValueCell;
    let explanationColumns;
    let valuePerUseColumns;
    let toggleButton;
    let isExplanationVisible = false;

    function formatCurrency(value) {
        return "$" + value.toLocaleString('en-US', {minimumFractionDigits: 2, maximumFractionDigits: 2});
        }

        function formatNumberWithCommas(num) {
        return num.toLocaleString('en-US');
    }

    function parseNumberFromInput(value) {
        // Remove commas and parse as integer
        return parseInt(value.replace(/,/g, '')) || 0;
    }

    function formatInputValue(input) {
        const rawValue = input.value.replace(/,/g, '');
        const numValue = parseInt(rawValue) || 0;
        if (numValue > 0) {
            input.value = formatNumberWithCommas(numValue);
        } else {
            input.value = '';
        }
    }

    // Update the calculate function:
    function calculate() {
        let total = 0;
        
        inputs.forEach(input => {
            const quantity = parseNumberFromInput(input.value);
            const product = input.getAttribute('data-product');
            const unitValue = productValues[product] || 0;
            const productTotal = quantity * unitValue;
            
            // Update the corresponding value cell
            const valueCell = input.closest('tr').querySelector('.value-cell');
            valueCell.textContent = formatCurrency(productTotal);
            
            total += productTotal;
        });
        
        // Update total
        totalValueCell.textContent = formatCurrency(total);
    }

    // Update the loadFiscalYear function to get the current raw value:
    function loadFiscalYear(year) {
        const data = fiscalYearData[year];
        if (data) {
            inputs.forEach(input => {
                const product = input.getAttribute('data-product');
                if (data.hasOwnProperty(product)) {
                    const currentValue = parseNumberFromInput(input.value);
                    animateValue(input, currentValue, data[product], 1500);
                }
            });
        }
    }

    // Update the animateValue function:
    function animateValue(input, start, end, duration) {
        const startTime = performance.now();
        
        function update() {
            const elapsed = performance.now() - startTime;
            const progress = Math.min(elapsed / duration, 1);
            
            const easeOut = 1 - Math.pow(1 - progress, 3);
            const current = Math.round(start + (end - start) * easeOut);
            
            input.value = formatNumberWithCommas(current);
            calculate();
            
            if (progress < 1) {
                requestAnimationFrame(update);
            }
        }
        
        requestAnimationFrame(update);
    }

    // Update the clearAll function:
    function clearAll() {
        inputs.forEach(input => {
            input.value = '';
        });
        calculate();
    }

    function toggleExplanation() {
        isExplanationVisible = !isExplanationVisible;

        // Toggle visibility of explanation columns
        explanationColumns.forEach(column => {
            column.classList.toggle('visible', isExplanationVisible);
        });

        // Toggle visibility of value per use columns
        valuePerUseColumns.forEach(column => {
            column.classList.toggle('visible', isExplanationVisible);
        });

        // Adjust colspan dynamically
        const totalRowFirstCell = document.querySelector('tfoot th:first-child');
        if (isExplanationVisible) {
            // When both explanation columns are visible: Quantity + Service + Values Explained = 3 columns
            totalRowFirstCell.setAttribute('colspan', '3');
        } else {
            // When explanation columns are hidden: Quantity + Service = 2 columns
            totalRowFirstCell.setAttribute('colspan', '2');
        }

        // Update button text
        toggleButton.textContent = isExplanationVisible
            ? 'Hide How Service Values Were Calculated'
            : 'Show How Service Values Were Calculated';
    }


    // Update the initializeCalculator function to add input event handlers:
    function initializeCalculator() {
        inputs = document.querySelectorAll('input[data-product]');
        totalValueCell = document.getElementById('total-value');
        explanationColumns = document.querySelectorAll('.values-explained-column');
        valuePerUseColumns = document.querySelectorAll('.value-per-use-column');
        toggleButton = document.querySelector('.toggle-explanation-button');

        if (!inputs.length || !totalValueCell || !toggleButton) {
            setTimeout(initializeCalculator, 100);
            return;
        }

        // Attach listeners to input fields
        inputs.forEach(input => {
            // Handle input formatting on blur (when user finishes typing)
            input.addEventListener('blur', function() {
                formatInputValue(this);
            });
            
            // Handle calculation on input
            input.addEventListener('input', function() {
                calculate();
            });
            
            // Handle focus to remove commas for easier editing
            input.addEventListener('focus', function() {
                const rawValue = this.value.replace(/,/g, '');
                if (rawValue === '0') {
                    this.value = '';
                } else {
                    this.value = rawValue;
                }
            });
        });

        // Attach listeners to fiscal year buttons
        const fyButtons = document.querySelectorAll('.fy-button');
        fyButtons.forEach(button => {
            const buttonText = button.textContent.trim();
            button.addEventListener('click', function() {
                loadFiscalYear(buttonText);
            });
        });

        // Attach listener to clear button
        const clearButton = document.querySelector('.clear-button');
        if (clearButton) {
            clearButton.addEventListener('click', clearAll);
        }

        // Attach listener to toggle explanation button
        toggleButton.addEventListener('click', toggleExplanation);

        // Initial calculation
        calculate();
    }

    // Initialize when ready
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', initializeCalculator);
    } else {
        initializeCalculator();
    }

    window.addEventListener('load', function() {
        if (!inputs || !inputs.length) {
            initializeCalculator();
        }
    });
</script>
'''
        return CODE_script
    def generate_style_and_script():
        Instructions = '''
<!-- 
HOW TO ADD THIS CODE
1) Open "Title & properties"
2) Add the Code to "HTML included in the <head> tag (except title and description)" 
-->

'''
        file_name = "style_and_script.html"
        with open(file_name, "w", encoding="utf-8") as f:
            f.write(Instructions)
            f.write(get_style())
            f.write(get_script())
            f.close()
    
    def get_table():
        CODE_table = '''
<table>
    <thead>
        <tr>
            <th>Quantity</th>
            <th data-role="resizable">Service</th>
            <th class="values-explained-column">Values Explained</th>
            <th class="value-per-use-column">Value Per Use ($)</th>
            <th>Value ($)</th>
        </tr>
    </thead>
    <tbody>
'''

        for service, service_info in Calculator_Settings_Dictionary.items():
            service_value = service_info[0]
            value_explained = service_info[1]
            CODE_table += "\t\t<tr>\n"
            CODE_table += f"\t\t\t<td><input data-product=\"{service}\" value=\"0\" min=\"0\" type=\"text\" /></td>\n"
            CODE_table += f"\t\t\t<td>{service}</td>\n"
            CODE_table += f"\t\t\t<td class=\"values-explained-column values-explained\">{value_explained}</td>\n"
            CODE_table += f'\t\t\t<td class="value-per-use-column value-per-use-cell">${service_value:.2f}</td>\n'
            CODE_table += "\t\t\t<td class=\"value-cell\">$0.00</td>\n"
            CODE_table += "\t\t</tr>\n"

        CODE_table += '''
    </tbody>
    <tfoot>
        <tr>
            <th colspan="2">Total Value Received</th>
            <th class="value-per-use-column"></th>
            <th id="total-value">$0.00</th>
        </tr>
    </tfoot>
</table>
'''
        return CODE_table
    def get_buttons():
        CODE_buttons = '''
<div class="buttons-container">
    <h2>Load Fiscal Year Data</h2>
    
    <h3>These values represent the total annual usage across all branches, not monthly personal usage.</h3>
        <div class="fy-buttons-row">
'''
        for FY, FY_info in Fiscal_Year_Dictionary.items():
            # Check if any service has a total of 0
            if any(total == 0 for service, total in FY_info):
                continue  # Skip this FY

            CODE_buttons += f'\t\t\t<button class="fy-button" type="button">{FY}</button>\n'
        
        CODE_buttons += '''
        </div>
    <div class="control-buttons-row"><button class="clear-button" type="button">Clear All</button><button class="toggle-explanation-button" type="button">Show How Service Values Were Calculated</button></div>
</div>
'''
        return CODE_buttons
    def generate_content_box_code():
        file_name = "content_box_code.html"
        with open(file_name, "w", encoding="utf-8") as f:
            f.write(get_table())
            f.write(get_buttons())
            f.close()
    
    def generate_prototype():
        file_name = "prototype.html"
        with open(file_name, "w", encoding="utf-8") as f:
            f.write(get_style())
            f.write(get_script())
            f.write(get_table())
            f.write(get_buttons())
            f.close()
    
    # Step 5) Create the Style_and_Script.html which will be added to the Head Tag:
    # Open "Title & properties"    ->   Add the Code to "HTML included in the <head> tag (except title and description)"
    generate_style_and_script()
    
    # Step 6) Create the content_box_code.html which will replace the code inside the content box
    generate_content_box_code()

    # Bonus) Create Prototype.html to quickly test the code in preview.
    generate_prototype()
    
import time
if __name__ == "__main__":
    start_time = time.time()
    generate_templates()
    end_time = time.time()
    print(f"Finished in {end_time - start_time:.3f} seconds.")
    input("Press Enter to Close the Program:")
