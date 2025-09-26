from openpyxl import load_workbook
import json
import re
import os

class TemplateBasedGenerator:
    def __init__(self, excel_file_path, templates_dir="templates"):
        self.excel_file_path = excel_file_path
        self.templates_dir = templates_dir
        self.workbook = load_workbook(excel_file_path)
        
        # Service mapping - maps row numbers to service keys
        self.service_mapping = {
            2: "books",
            7: "dvds", 
            11: "audiobooks",
            12: "ebooks",
            22: "meeting-room",
            23: "study-room",
            25: "archive-lab",
            27: "wifi",
            29: "online-learning",
            31: "research",
            33: "technical-instructions-and-training",
            35: "computer-usage",
            41: "program-child",
            42: "program-teen",
            45: "program-adult"
        }
        
        # Service display names
        self.service_names = {
            "books": "Books Borrowed per Month",
            "dvds": "DVDs Borrowed per Month",
            "audiobooks": "Audiobooks Borrowed per Month", 
            "ebooks": "eBooks Borrowed per Month",
            "meeting-room": "Meeting Room Usage Hours per Month",
            "study-room": "Study Room Usage Hours per Month",
            "archive-lab": "Archive Lab Usage Hours per Month",
            "wifi": "WiFi Use Hours per Month",
            "online-learning": "Online Learning Users per Month",
            "research": "Research Questions Asked per Month",
            "technical-instructions-and-training": "Technical Instructions and Training Hours per Month",
            "computer-usage": "Computer Use Hours per Month",
            "program-child": "Program/Class Attended per Month - Child",
            "program-teen": "Program/Class Attended per Month - Teen",
            "program-adult": "Program/Class Attended per Month - Adult"
        }

    def load_template(self, template_name):
        """Load a template file"""
        template_path = os.path.join(self.templates_dir, template_name)
        if not os.path.exists(template_path):
            # Try current directory
            template_path = template_name
        
        try:
            with open(template_path, 'r', encoding='utf-8') as f:
                return f.read()
        except FileNotFoundError:
            print(f"Template file not found: {template_path}")
            return None

    def extract_fiscal_year_data(self):
        """Extract fiscal year data from the 'Fiscal Year Data' worksheet"""
        ws = self.workbook['Fiscal Year Data']
        
        # Find fiscal year columns by scanning the header row
        fiscal_years = {}
        header_row = 1
        
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(header_row, col).value
            if header_value and isinstance(header_value, str):
                # Look for pattern like "FY 2022-23" or similar
                fy_match = re.search(r'FY\s*(\d{4})-(\d{2})', header_value)
                if fy_match:
                    year_key = f"{fy_match.group(1)}-{fy_match.group(2)}"
                    fiscal_years[year_key] = col
        
        # Extract data for each fiscal year and service
        fiscal_data = {}
        for year, col in fiscal_years.items():
            fiscal_data[year] = {}
            for row_num, service_key in self.service_mapping.items():
                cell_value = ws.cell(row_num, col).value
                # Convert to integer, default to 0 if empty or invalid
                try:
                    fiscal_data[year][service_key] = int(cell_value) if cell_value is not None else 0
                except (ValueError, TypeError):
                    fiscal_data[year][service_key] = 0
        
        return fiscal_data

    def extract_values_per_service(self):
        """Extract service values and explanations from 'Values Per Service' worksheet"""
        ws = self.workbook['Values Per Service']
        
        service_values = {}
        service_explanations = {}
        
        # Scan the worksheet to find service data
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row, col).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    
                    # Try to match service names
                    for service_key, service_name in self.service_names.items():
                        if service_name.lower() in cell_str or service_key.replace('-', ' ') in cell_str:
                            # Look for value and explanation in adjacent columns
                            for check_col in range(col + 1, min(col + 4, ws.max_column + 1)):
                                check_value = ws.cell(row, check_col).value
                                if check_value:
                                    try:
                                        # Try to parse as currency/float
                                        if isinstance(check_value, (int, float)):
                                            value = float(check_value)
                                            service_values[service_key] = value
                                        elif isinstance(check_value, str):
                                            # Try to extract number from string
                                            value_match = re.search(r'[\$]?(\d+\.?\d*)', check_value)
                                            if value_match:
                                                service_values[service_key] = float(value_match.group(1))
                                            elif len(check_value) > 20:  # Likely an explanation
                                                service_explanations[service_key] = check_value.strip()
                                    except (ValueError, TypeError):
                                        if isinstance(check_value, str) and len(check_value) > 20:
                                            service_explanations[service_key] = check_value.strip()
                            break
        
        # Fill in any missing values with defaults
        default_values = {
            "books": 16.00, "dvds": 10.00, "audiobooks": 25.00, "ebooks": 6.00,
            "meeting-room": 30.00, "study-room": 10.00, "archive-lab": 45.00, "wifi": 4.00,
            "online-learning": 75.00, "research": 15.00, "technical-instructions-and-training": 25.00,
            "computer-usage": 12.00, "program-child": 10.00, "program-teen": 13.00, "program-adult": 17.50
        }
        
        default_explanations = {
            "books": "Average retail price of a new paperback book.",
            "dvds": "Matches typical purchase cost for a new or recent DVD.",
            "audiobooks": "Reflects average price for full-length audiobooks from Audible or publishers.",
            "ebooks": "Reflects average cost of popular consumer eBooks (not free/public domain).",
            "meeting-room": "Based on hourly rental rates of comparable private meeting spaces in community centers.",
            "study-room": "Aligns with value of private study rooms or co-working quiet spaces per hour.",
            "archive-lab": "Based on commercial digitization services (VHS transfer, slide scanning, etc.) including staff support.",
            "wifi": "Reflects value of secure high-speed internet access in public/commercial spaces.",
            "online-learning": "Represents estimated monthly retail value of bundled premium services (e.g., LinkedIn Learning, Mango, Morningstar, PressReader, ProQuest).",
            "research": "Equivalent to hiring a research assistant or paying for a reference consultation.",
            "technical-instructions-and-training": "Based on average hourly rate for adult education, digital literacy classes, or technology tutoring through community centers, workforce training, or commercial platforms.",
            "computer-usage": "Matches internet café/computer rental rates per hour, especially with MS Office access.",
            "program-child": "Estimated value based on costs for enrichment or storytime programs in community centers.",
            "program-teen": "Based on creative or educational teen-focused activities (e.g. coding, crafts).",
            "program-adult": "Reflects value of workshops, lectures, or adult education programs elsewhere."
        }
        
        # Fill in missing values
        for service_key in self.service_names.keys():
            if service_key not in service_values:
                service_values[service_key] = default_values[service_key]
            if service_key not in service_explanations:
                service_explanations[service_key] = default_explanations[service_key]
        
        return service_values, service_explanations

    def generate_table_rows(self, service_values, service_explanations):
        """Generate HTML table rows"""
        rows = []
        for service_key in self.service_names.keys():
            service_name = self.service_names[service_key]
            value = service_values.get(service_key, 0.00)
            explanation = service_explanations.get(service_key, "")
            
            row = f'''                <tr>
                    <td><input data-product="{service_key}" value="0" min="0" type="text" /></td>
                    <td>{service_name}</td>
                    <td class="values-explained-column values-explained">{explanation}</td>
                    <td class="value-per-use-column value-per-use-cell">${value:.2f}</td>
                    <td class="value-cell">$0.00</td>
                </tr>'''
            rows.append(row)
        
        return "\n".join(rows)

    def generate_fiscal_year_buttons(self, fiscal_data):
        """Generate fiscal year buttons HTML"""
        buttons = []
        for year in sorted(fiscal_data.keys()):
            buttons.append(f'<button class="fy-button" type="button">FY {year}</button>')
        return "".join(buttons)

    def generate_product_values_js(self, service_values):
        """Generate JavaScript product values object"""
        js_values = []
        for service_key, value in service_values.items():
            js_values.append(f'        "{service_key}": {value:.2f}')
        return ",\n".join(js_values)

    def generate_fiscal_year_data_js(self, fiscal_data):
        """Generate JavaScript fiscal year data object"""
        js_fiscal_data = []
        for year, year_data in fiscal_data.items():
            js_year_services = []
            for service_key, count in year_data.items():
                js_year_services.append(f'            "{service_key}": {count}')
            
            year_block = f'        "{year}": {{\n{",\n".join(js_year_services)}\n        }}'
            js_fiscal_data.append(year_block)
        
        return ",\n".join(js_fiscal_data)

    def generate_button_logic_js(self, fiscal_data):
        """Generate JavaScript for fiscal year button logic"""
        logic_lines = []
        for year in sorted(fiscal_data.keys()):
            logic_lines.append(f'            if (buttonText.includes(\'{year}\')) year = \'{year}\';')
        return "\n".join(logic_lines)

    def generate_html_from_template(self, css_file="Style_and_Script.html"):
        """Generate HTML using templates"""
        # Extract data from Excel
        fiscal_data = self.extract_fiscal_year_data()
        service_values, service_explanations = self.extract_values_per_service()
        
        # Load CSS content
        css_content = ""
        if os.path.exists(css_file):
            with open(css_file, 'r', encoding='utf-8') as f:
                content = f.read()
                # Extract CSS from <style> tags
                css_match = re.search(r'<style>(.*?)</style>', content, re.DOTALL)
                if css_match:
                    css_content = css_match.group(1).strip()
        
        # Generate components
        table_rows = self.generate_table_rows(service_values, service_explanations)
        fy_buttons = self.generate_fiscal_year_buttons(fiscal_data)
        product_values_js = self.generate_product_values_js(service_values)
        fiscal_year_data_js = self.generate_fiscal_year_data_js(fiscal_data)
        button_logic_js = self.generate_button_logic_js(fiscal_data)
        
        # Load and populate HTML template
        html_template = self.load_template("template.html")
        if not html_template:
            # Inline template if file not found
            html_template = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Library Service Value Calculator</title>
    <style>
        {{CSS_CONTENT}}
    </style>
</head>
<body>
    <div class="container">
        <h1>Library Service Value Calculator</h1>
        <table>
            <thead>
                <tr>
                    <th>Quantity</th>
                    <th data-role="resizable">Service</th>
                    <th class="values-explained-column">Values Explained</th>
                    <th class="value-per-use-column">Value Per Use ($)</th>
                    <th>Value ($)</th>
                </tr>
            </thead>
            <tbody>
{{TABLE_ROWS}}
            </tbody>
            <tfoot>
                <tr>
                    <th colspan="2">Total Value Received</th>
                    <th class="value-per-use-column"></th>
                    <th id="total-value">$0.00</th>
                </tr>
            </tfoot>
        </table>
        <div class="buttons-container">
            <h2>Load Fiscal Year Data</h2>
            <div class="fy-buttons-row">{{FISCAL_YEAR_BUTTONS}}</div>
            <div class="control-buttons-row">
                <button class="clear-button" type="button">Clear All</button>
                <button class="toggle-explanation-button" type="button">Show How Service Values Were Calculated</button>
            </div>
        </div>
    </div>
    {{JAVASCRIPT_CONTENT}}
</body>
</html>'''
        
        # Load and populate JavaScript template
        js_template = self.load_template("template.js")
        if not js_template:
            # Inline template if file not found
            js_template = '''<script>
    const productValues = {
{{PRODUCT_VALUES}}
    };

    const fiscalYearData = {
{{FISCAL_YEAR_DATA}}
    };

    let inputs;
    let totalValueCell;
    let explanationColumns;
    let valuePerUseColumns;
    let toggleButton;
    let isExplanationVisible = false;

    function formatCurrency(value) {
        return "$" + value.toLocaleString('en-US', {minimumFractionDigits: 2, maximumFractionDigits: 2});
    }

    function formatNumberWithCommas(num) {
        return num.toLocaleString('en-US');
    }

    function parseNumberFromInput(value) {
        return parseInt(value.replace(/,/g, '')) || 0;
    }

    function formatInputValue(input) {
        const rawValue = input.value.replace(/,/g, '');
        const numValue = parseInt(rawValue) || 0;
        if (numValue > 0) {
            input.value = formatNumberWithCommas(numValue);
        } else {
            input.value = '';
        }
    }

    function calculate() {
        let total = 0;
        
        inputs.forEach(input => {
            const quantity = parseNumberFromInput(input.value);
            const product = input.getAttribute('data-product');
            const unitValue = productValues[product] || 0;
            const productTotal = quantity * unitValue;
            
            const valueCell = input.closest('tr').querySelector('.value-cell');
            valueCell.textContent = formatCurrency(productTotal);
            
            total += productTotal;
        });
        
        totalValueCell.textContent = formatCurrency(total);
    }

    function loadFiscalYear(year) {
        const data = fiscalYearData[year];
        if (data) {
            inputs.forEach(input => {
                const product = input.getAttribute('data-product');
                if (data.hasOwnProperty(product)) {
                    const currentValue = parseNumberFromInput(input.value);
                    animateValue(input, currentValue, data[product], 1500);
                }
            });
        }
    }

    function animateValue(input, start, end, duration) {
        const startTime = performance.now();
        
        function update() {
            const elapsed = performance.now() - startTime;
            const progress = Math.min(elapsed / duration, 1);
            
            const easeOut = 1 - Math.pow(1 - progress, 3);
            const current = Math.round(start + (end - start) * easeOut);
            
            input.value = formatNumberWithCommas(current);
            calculate();
            
            if (progress < 1) {
                requestAnimationFrame(update);
            }
        }
        
        requestAnimationFrame(update);
    }

    function clearAll() {
        inputs.forEach(input => {
            input.value = '';
        });
        calculate();
    }

    function toggleExplanation() {
        isExplanationVisible = !isExplanationVisible;

        explanationColumns.forEach(column => {
            column.classList.toggle('visible', isExplanationVisible);
        });

        valuePerUseColumns.forEach(column => {
            column.classList.toggle('visible', isExplanationVisible);
        });

        const totalRowFirstCell = document.querySelector('tfoot th:first-child');
        if (isExplanationVisible) {
            totalRowFirstCell.setAttribute('colspan', '3');
        } else {
            totalRowFirstCell.setAttribute('colspan', '2');
        }

        toggleButton.textContent = isExplanationVisible
            ? 'Hide How Service Values Were Calculated'
            : 'Show How Service Values Were Calculated';
    }

    function initializeCalculator() {
        inputs = document.querySelectorAll('input[data-product]');
        totalValueCell = document.getElementById('total-value');
        explanationColumns = document.querySelectorAll('.values-explained-column');
        valuePerUseColumns = document.querySelectorAll('.value-per-use-column');
        toggleButton = document.querySelector('.toggle-explanation-button');

        if (!inputs.length || !totalValueCell || !toggleButton) {
            setTimeout(initializeCalculator, 100);
            return;
        }

        inputs.forEach(input => {
            input.addEventListener('blur', function() {
                formatInputValue(this);
            });
            
            input.addEventListener('input', function() {
                calculate();
            });
            
            input.addEventListener('focus', function() {
                const rawValue = this.value.replace(/,/g, '');
                if (rawValue === '0') {
                    this.value = '';
                } else {
                    this.value = rawValue;
                }
            });
        });

        const fyButtons = document.querySelectorAll('.fy-button');
        fyButtons.forEach(button => {
            const buttonText = button.textContent.trim();
            let year = '';
{{FISCAL_YEAR_BUTTON_LOGIC}}
            
            if (year) {
                button.addEventListener('click', function() {
                    loadFiscalYear(year);
                });
            }
        });

        const clearButton = document.querySelector('.clear-button');
        if (clearButton) {
            clearButton.addEventListener('click', clearAll);
        }

        toggleButton.addEventListener('click', toggleExplanation);

        calculate();
    }

    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', initializeCalculator);
    } else {
        initializeCalculator();
    }

    window.addEventListener('load', function() {
        if (!inputs || !inputs.length) {
            initializeCalculator();
        }
    });
</script>'''
        
        # Populate JavaScript template
        js_content = js_template.replace('{{PRODUCT_VALUES}}', product_values_js)
        js_content = js_content.replace('{{FISCAL_YEAR_DATA}}', fiscal_year_data_js)
        js_content = js_content.replace('{{FISCAL_YEAR_BUTTON_LOGIC}}', button_logic_js)
        
        # Populate HTML template
        html_content = html_template.replace('{{CSS_CONTENT}}', css_content)
        html_content = html_content.replace('{{TABLE_ROWS}}', table_rows)
        html_content = html_content.replace('{{FISCAL_YEAR_BUTTONS}}', fy_buttons)
        html_content = html_content.replace('{{JAVASCRIPT_CONTENT}}', js_content)
        
        return html_content

    def save_generated_files(self, output_file="generated_calculator.html"):
        """Generate and save the HTML file"""
        try:
            html_content = self.generate_html_from_template()
            
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            print(f"✅ HTML file generated successfully: {output_file}")
            
            # Print summary
            fiscal_data = self.extract_fiscal_year_data()
            service_values, _ = self.extract_values_per_service()
            
            print(f"📊 Data Summary:")
            print(f"   • Found {len(fiscal_data)} fiscal years: {', '.join(sorted(fiscal_data.keys()))}")
            print(f"   • Found {len(service_values)} services with values")
            print(f"   • Total data points: {sum(len(data) for data in fiscal_data.values())}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error generating file: {str(e)}")
            return False

# Example usage and main execution
if __name__ == "__main__":
    import sys
    
    # Default Excel file path
    excel_path = "CalculatorSettings.xlsx"
    
    # Check if Excel file path provided as command line argument
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    
    # Check if file exists
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        print("Please make sure the file exists and the path is correct.")
        print("Usage: python template_generator.py [path_to_excel_file]")
        sys.exit(1)
    
    try:
        print(f"🔄 Processing Excel file: {excel_path}")
        generator = TemplateBasedGenerator(excel_path)
        
        # Generate and save the HTML file
        success = generator.save_generated_files()
        
        if success:
            print("🎉 Generation complete!")
        else:
            print("❌ Generation failed!")
            sys.exit(1)
            
    except Exception as e:
        print(f"❌ Unexpected error: {str(e)}")
        sys.exit(1)