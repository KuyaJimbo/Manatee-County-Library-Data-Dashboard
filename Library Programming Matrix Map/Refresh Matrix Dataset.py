import csv, os, re
from time import time
# Import Helper Functions From Python Libraries
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def classify_event_program_by_title(wb, EventID, Title, Categories):
    match_found = False
    ws = wb["EventProgram"]
    Title = Title.lower()  # Case-insensitive match
    if "cancelled" in Title or "canceled" in Title:
        return

    Categories = [cat.lower() for cat in Categories]

    # ---------- STORY TIME Programs ----------
    # Define keyword patterns
    title_keywords = ["time", "40 carrots", "Puppet", "Dr. Seuss Day", "Soar"]
    category_keywords = ["literacy > storytimes"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Story Time"])
        match_found = True


    # ---------- MAKERSPACE + CRAFTS Programs ----------
    # Define keyword patterns
    title_keywords = [
        "Hogwarts", "Sleeping Mat", "Scrapbook", "Sensory Sensitive", "Spring Festival",
        "Stitch", "Ginger", "Outdoors", "Potion", "Sand Dollar",
        "Yarn", "Lanterns", "Mirrors", "Pencil", "Wreaths",
        "Oyster", "Knit", "Ornaments", "Card-making", "3D Print",
        "Art Lab", "Coloring Club", "Drop", "Model Magic", "Art Media",
        "Art Station", "Tea with an Artist", "Sew", "Valentines", "Art House",
        "Art with Leaves", "Mood Board", "t-shirts", "Bad Art", "Making Cards",
        "Bookmarks", "Bedazzle", "Quilt", "TinkerCAD", "Wind Chimes",
        "Book Folding", "Bookmark Challenge", "Bottle Cap", "Writers Group", "Builders Club",
        "Calligraphy", "Canvas", "Cardmaking", "Chalk Art", "Collage",
        "Painting", "Craft", "Hydrangeas", "Create", "Creative",
        "Design", "Crochet", "Decorate", "Paper", "DIY",
        "Art Studio", "Drawing", "Dream Catchers", "Teacups", "Photo",
        "Makerspace", "Resin", "FOL", "Jar", "Gift Wrap",
        "Champagne Flutes", "Glowforge", "Gnome", "Terrarium", "Harry Potter",
        "Art Corner", "Workshop", "Star Wars", "Superhero", "Polymer",
        "Colors", "Art Club", "Origami", "Cricut", "Seashell",
        "Seaside Quilters", "Write Away!", "Jelly Bean", "Poet", "Magic Crystals",
        "Bandana", "Headband!", "Patchworkers", "Memoir", "Memory Makers",
        "Bottle Decoding", "Bracelet", "New Art Mediums", "Journaling", "Writing"
    ]
    title_exceptions = ["plant", "weight", "conservation"]
    category_keywords = ["makerspace", "crafts", "Drawing", "Virtual Reality"]
    # Search for keywords
    if (any(keyword.lower() in Title for keyword in title_keywords) and not any(keyword.lower() in Title for keyword in title_exceptions)) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Makerspace and Workshop"])
        match_found = True
    
    # ---------- ONE ON ONE TECH SUPPORT Programs ----------
    # Define keyword patterns
    title_keywords = [
        "Apple Products", "Digitizing", "Virtual Reality", "Excel", "iPhone",
        "Tech", "App", "iPad", "Microsoft", "Library 101", "Computers", 
        "Archive Lab", "Podcasting", "Libby", "Scratch"
    ]
    category_keywords = ["Technology", "Technology > Computers", "Technology > iPhone/iPad", 
                         "Technology > One-on-One Help", "Technology > STEM"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Tech Support"])
        match_found = True


    # ---------- BOOK CLUBS Programs ----------
    # Define keyword patterns
    title_keywords = [
        "Discussion", "Book Circle", "Book Club", "Keep your secrets", "Literary Travel",
        "Improv", "Mystery", "Graphic Novels", "Author Fair"
    ]
    category_keywords = ["Literacy > Book Clubs & Author Talks"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Book Club"])
        match_found = True

    # ---------- READER'S ADVISORY Programs ----------
    category_keywords = ["Reader's Advisory"]
    # Search for keywords
    if any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Reader's Advisory"])
        match_found = True
    

    # ---------- Discovery Center Programs ----------
    # Define keyword patterns
    title_keywords = [
        "Moogician", "Stev", "Polly", "Mangrove", "Razzmatazz",
        "Lawn for Tweens", "Roar", "Jungle Gardens", "Campbell", "Adventure Club",
        "Kids Club", "After School Chill", "Anime Club", "Privateers", "Read-Aloud",
        "Parrot Show", "Block Fest", "Reads-A-Lot", "Cephalopalooza", "Colorful",
        "Curiosity Club", "Didgeridoo", "Dog", "obstacle course", "Smokey",
        "Summer Learning", "Traveler's", "Mad Science", "Out of My Hands", "School of Rock",
        "Showtime", "STEAM", "Mart", "JiggleMan!", "Balloon Show",
        "Make Believe"
    ]
    category_keywords = ["Summer Learning"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Discovery Center"])
        match_found = True


    # ---------- GENEALOGY SERVICES Programs ----------
    # Define keyword patterns
    title_keywords = ["Heritage","Genealogists","Genealogical","Genealogy","Family History","DNA"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords):
        ws.append([EventID, "Genealogy Services"])
        match_found = True


    # ---------- LANGUAGE AND CULTURE Programs ----------
    # Define keyword patterns
    title_keywords = [
        "Sign Language", "ESL", "ESOL", "French", "Spanish Class", "Mandarin",
        "Conversations","Culture","Friends 2025 Lecture & Travel","Morocc","Atlas"
    ]
    category_keywords = ["Language Learning", "International", "Travel & Leisure", "Voyagers"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Language and Culture"])
        match_found = True


    # ---------- LOCAL HISTORY AND ARCHIVES Programs ----------
    # Define keyword patterns
    title_keywords = [
        "Pirates of the Florida Coast", "American Revolution", "Ghosts", "Palma Sola", "Ship Passenger Records",
        "Presidency", "Hamilton", "History", "Military Challenge Coins", "Veteran",
        "War", "Leading Ladies", "Steel Ring Academy", "Memorial", "Flags",
        "Museum", "Frank Lloyd Wright", "Florida Authors", "Ringling", "Spooky"
    ]
    category_keywords = ["FLORIDA"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Local History & Archives"])
        match_found = True


    # ---------- FITNESS AND WELLNESS Programs ----------
    # Define keyword patterns
    title_keywords = [
        "Reset and Renew", "Peer Support", "Weight", "Nourish", "Healing",
        "Stroller Strides", "Swag", "After School Chill", "Aerobics", "Wellness",
        "Pilates", "Alzheimer", "Cooking", "Dancing", "Cancer",
        "Yoga", "Zumba", "Winemaking", "Blood", "Safety",
        "Feast", "Prevention", "Health Resource", "Foods That Heal", "Screening",
        "HIV", "Mental Health", "Balance", "Healthiest Self", "Hike",
        "Zen Zone", "Tai Chi", "Sound Bath", "Public Health", "Nutrition",
        "Illness", "Meditation", "Herbal", "Cuisine", "Jazzercise",
        "Health Department", "Medicare", "Senior Health", "Senior Living", "SHINE",
        "Dance Along", "Memory Cafe", "Mindful", "Dance-A-Long"
    ]
    category_keywords = ["Health/Fitness/Wellness", "Sports & Leisure"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Fitness and Wellness"])
        match_found = True

    # ---------- ENTERTAINMENT Programs ----------
    title_keywords = [
        "Pinochle", "Morning Card-toons", "Wonder Lab", "Wonderland", "Dance Mob",
        "SOM", "Chocolate Factory", "Stuffed Animal", "Sweet", "Taskmaster",
        "Y2K", "Thanksgiving", "Beauty and the Beast", "Anniversary", "Open Mic",
        "Parrish Playworks", "Peter Rabbit", "Pigeon Party", "Polo Club", "St. Patrick",
        "OZ", "Novemberfest", "Candy Land", "Bingo", "Board Game Bash",
        "Boggle", "Bricks4Kidz", "Bridge", "Bubble Party", "Chess",
        "Juguemos", "Dungeons", "Duplo", "Egg Hunt", "End of Summer",
        "DND", "The Office", "Fall Frienzy", "Fright", "Movie",
        "Fantasy Map", "Foam", "Fort Night", "Game", "Galentine's Night",
        "Holiday Party", "Geocaching", "Garden Party", "Parade", "Santa",
        "Kickoff", "Kick-Off", "Smash", "RPG", "Pokemon",
        "Pokémon", "Peeps", "Pizza", "Nintendo", "Minecraft",
        "Mario", "jong", "Lego", "Just Dance", "Holidays Around the World",
        "Playground", "Playgroup", "Golf", "Oreo", "Hunt",
        "Gaming", "Waffle party", "Playdough", "Trivia", "Candyland",
        "Trick or Treat", "Truck", "Polar Express", "Lil' Manatee Cove", "Lo-Fi"
    ]
    category_keywords = ["Fun & Games", "Arts & Entertainment", "Concert"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Entertainment"])
        match_found = True
    
    # ---------- NATURE AND HOME Programs ----------
    # Define keyword patterns
    title_keywords = [
        "Marine Lab", "Family Treasures", "Yard Drinking", "Water Conservation", "Recycling",
        "Turtle Watch", "Gardener", "Irrigation", "Build Homes", "Garden",
        "Dipnetting", "Decluttering", "Home Decor", "Landscapes", "Recycle",
        "Agriculture", "Natural Resources", "Seed", "Wildlife", "Plant",
    ]
    title_exceptions = ["Teacups"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) and not any(keyword.lower() in Title for keyword in title_exceptions):
        ws.append([EventID, "Nature and Home"])
        match_found = True

    # ---------- MUSIC AND FILM Programs ----------
    # Define keyword patterns
    title_keywords = ["Ukulele","Broadway","Concert","Blues","Cinema","Guitar","Opera","Instrument","Music"]
    category_keywords = ["Music Instruction"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords) or any(keyword.lower() in Categories for keyword in category_keywords):
        ws.append([EventID, "Music and Film"])
        match_found = True

    # --------- IMPORTANT MEETINGS ---------
    # Define keyword patterns
    title_keywords = ["Board Meeting","Teen Advisory Board","Town Hall","Literacy Council"]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords):
        ws.append([EventID, "Important Meeting"])
        match_found = True

    # --------- LIFE SKILLS AND COMMUNITY RESOURCES MEETINGS ---------
    # Define keyword patterns
    title_keywords =  [
        "Welcome to Our World", "Families Educational Seminar", "Tax", "Career", "Financial",
        "Finanzas", "Job Fair", "Scams", "FAFSA", "Antiques", "Food Bank", "Vote", "Sale", 
        "Community", "Alliance Gives Back"
    ]
    # Search for keywords
    if any(keyword.lower() in Title for keyword in title_keywords):
        ws.append([EventID, "Life Skills and Community Resource"])
        match_found = True
    
    # ---------- UNMATCHED PROGRAMS ----------
    if not match_found:
        ws.append([EventID, "Unmatched Events"])

def convert_time_to_hours(time_str):
    """Convert a time string in 'HH:MM' format to hours as a float."""
    if not time_str:
        return -1
    try:
        hours, minutes = map(int, time_str.split(":"))
        return hours + minutes / 60
    except ValueError:
        return -1

def format_hours(raw_hours):
    hours = int(raw_hours)
    minutes = int(round((raw_hours - hours) * 60))
    return f"{hours} hours and {minutes} min"

def calculate_duration(start_time, end_time):
    """
    Calculate the duration between two time strings in hours.
    Returns -1 if the start_time or end_time is invalid.
    """
    start_hours = convert_time_to_hours(start_time)
    end_hours = convert_time_to_hours(end_time)
    if start_hours == -1 or end_hours == -1:
        return -1
    duration = end_hours - start_hours
    return duration

def calculate_event_staff_time(start_time, end_time, set_up_time, tear_down_time):
    """
    Calculate the total staff time for an event.
    Returns at minimum the event duration + any additional prep or post event time.
    If returns 0, the event is considered invalid.
    """
    # Event Duration) Calculate time between start_time and end_time
    event_duration = calculate_duration(start_time, end_time)
    # if event_duration is invalid, it means there was no listed event time
    if event_duration == -1:
        event_duration = 0

    # Prep Time) Calculate time between set_up_time and start time
    set_up_duration = calculate_duration(set_up_time, start_time)
    # if set_up_duration is invalid, it means there was no listed set up time
    if set_up_duration == -1:
        set_up_duration = 0

    # Post Time) Calculate time between end_time and tear_down_time
    tear_down_duration = calculate_duration(end_time, tear_down_time)
    # if tear_down_duration is invalid, it means there was no listed tear down time
    if tear_down_duration == -1:
        tear_down_duration = 0

    total_staff_time = event_duration + set_up_duration + tear_down_duration
    return total_staff_time

def create_dataset_template():
    # Create a Workbook
    wb = Workbook()

    # Give the Workbook The Following Worksheets
    workbook_setup = {  "EventInformation": [ "EventID", "Title", "Description", "Location", "Library Branch",
                                            "Event Organizer", "Presenter", "Audiences", 
                                            "Categories", "Internal Tags"],
                        "EventAudiences": ["EventID", "Audience"],
                        "EventCategories": ["EventID", "Category"],
                        "EventInternalTags": ["EventID", "Internal Tag"],
                        "EventTimes": ["EventID", "Event Start Date", "Event End Date", "All Day Event",
                                    "Start Time", "End Time", "Set Up Time", "Tear Down Time", 
                                    "Duration", "Staff Time", "Formatted Duration", "Different Dates"],
                        "EventParticipation": ["EventID", "Registration Required", "In-Person Seats", 
                                                "Online Seats", "Confirmed Registrations", 
                                                "Waiting-List Registrations", "Cancelled Registrations", 
                                                "Anticipated Attendance", "Actual Attendance (In-Person)", 
                                                "Actual Attendance (Online)", "Confirmed Attendance"],
                        "EventProgram": ["EventID", "Program Type"]
                     }

    # remove the first sheet
    wb.remove(wb["Sheet"])

    # Add Worksheets and Headers
    for sheet_name, headers in workbook_setup.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
    
    return wb

def create_matrix_map_settings():
    # Create a Workbook
    wb = Workbook()

    # Give the Workbook The Following Worksheets (avg)
    workbook_setup = {  "Program Options": [ "Program Type", "Cost Score", "Impact Score", 
                                                      "Connection and Belonging", "Trust", "Access", "Community Reach", "Creativity and Joy", "Strategic Fit", "Recommended Impact Score", 
                                                      "Cost Recovery / Net Cost", "Resource Intensity", "Scaling Difficulty", "Funding Dependency", "Operational Complexity", "Recommended Cost Score"],

                        "Category Options": ["Category", "Cost Score", "Impact Score", 
                                             "Connection and Belonging", "Trust", "Access", "Community Reach", "Creativity and Joy", "Strategic Fit", "Recommended Impact Score", 
                                             "Cost Recovery / Net Cost", "Resource Intensity", "Scaling Difficulty", "Funding Dependency", "Operational Complexity", "Recommended Cost Score"],

                        "Audience Options" : ["Audience", "Cost Score", "Impact Score", 
                                              "Connection and Belonging", "Trust", "Access", "Community Reach", "Creativity and Joy", "Strategic Fit", "Recommended Impact Score", 
                                              "Cost Recovery / Net Cost", "Resource Intensity", "Scaling Difficulty", "Funding Dependency", "Operational Complexity", "Recommended Cost Score"],
                     }

    # remove the first sheet
    wb.remove(wb["Sheet"])

    # Add Worksheets and Headers
    for sheet_name, headers in workbook_setup.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
    
    return wb

def find_data_source_file():
    # Search for the CSV file in the current directory
    # Pattern: starts with "lc_events_"
    csv_files = [f for f in os.listdir('.') if re.match(r'lc_events_.*\.csv$', f)]
    if not csv_files:
        print("No lc_events_*.csv file found in the current directory.")
        print("Could not update Matrix Map Dataset.")
        exit(1)
    else:
        print(f"Found data source file: {csv_files[0]}")
    return csv_files[0]

def read_csv_and_populate_workbook(wb, source_file):
    LocationSet = set()
    CategorySet = set()
    AudienceSet = set()
    InternalTagSet = set()
    BranchSet = set()
    FiscalYearSet = set()

    with open(source_file, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            # Skip header row
            if reader.line_num == 1:
                continue

            EventID = row[0]
            Title = row[1]

            # Skip events that were cancelled
            if "cancel" in Title.lower():
                continue
            
            Description = row[2] #*Not included to reduce unnecessary dataset size
            EventStartDate = row[3]

            Year = EventStartDate.split("-")[0]
            FiscalYear = "FY " + Year + "-" + str((int(Year) + 1) % 100)
            FiscalYearSet.add(FiscalYear)

            EventEndDate = row[4] # *Not included to reduce unnecessary dataset size

            DifferentDates = (EventStartDate != EventEndDate)
            
            StartTime = row[5]      # Always Required
            EndTime = row[6]        # Sometimes Blank, if StartTime is "All Day Event"
            SetUpTime = row[7]      # Sometimes Blank
            TearDownTime = row[8]   # Sometimes Blank

            AllDayEvent = (StartTime == "All Day Event")
            if AllDayEvent:
                StartTime = "00:00"

            Duration = calculate_duration(StartTime, EndTime) # Duration is -1 if no valid StartTime or EndTime
            StaffTime = calculate_event_staff_time(StartTime, EndTime, SetUpTime, TearDownTime)

            FormattedDuration = format_hours(Duration)

            Location = row[9]
            LocationSet.add(Location)

            LibraryBranch = row[10]
            BranchSet.add(LibraryBranch)

            EventOrganizer = row[11]
            Presenter = row[12]
            Audiences = row[13]
            AudiencesList = [audience.strip() for audience in Audiences.split(",")]

            Categories = row[14]
            CategoriesList = [category.strip() for category in Categories.split(",")]

            # PublishingStatus = row[15] *Always set to "Published" Not included to reduce unnecessary dataset size
            InternalTags = row[16]
            InternalTagsList = [tag.strip() for tag in InternalTags.split(",")]
            # EventNote = row[17] *Not included to reduce unnecessary dataset size
            RegistrationRequired = row[18]
            InPersonSeats = row[19]
            OnlineSeats = row[20]
            ConfirmedRegistrations = row[21]
            WaitlistRegistrations = row[22]
            CancelledRegistrations = row[23]
            AnticipatedAttendance = row[24]
            ActualAttendance_InPerson = row[25]
            ActualAttendance_Online = row[26]
            ConfirmedAttendance = row[27]
            # EventURL = row[28] *Not included to reduce unnecessary dataset size

            # Add to Event Information Table
            ws = wb["EventInformation"]
            ws.append([EventID, Title, Description, Location, LibraryBranch, EventOrganizer, Presenter, Audiences, Categories, InternalTags])

            # Add to Event Audiences Table
            ws = wb["EventAudiences"]
            for audience in AudiencesList:
                ws.append([EventID, audience])

            for audience in set(AudiencesList):
                if audience not in AudienceSet:
                    AudienceSet.add(audience)

            # Add to Event Categories Table
            ws = wb["EventCategories"]
            for category in CategoriesList:
                ws.append([EventID, category])

            for category in set(CategoriesList):
                if category not in CategorySet:
                    CategorySet.add(category)

            # Add to Event Internal Tags Table
            ws = wb["EventInternalTags"]
            for tag in InternalTagsList:
                ws.append([EventID, tag])

            for tag in set(InternalTagsList):
                if tag not in InternalTagSet:
                    InternalTagSet.add(tag)

            # Add to Event Times Table
            ws = wb["EventTimes"]
            ws.append([EventID, EventStartDate, EventEndDate, AllDayEvent, StartTime, EndTime, SetUpTime, TearDownTime, Duration, StaffTime, FormattedDuration, DifferentDates])

            # Add to Event Participation Table
            ws = wb["EventParticipation"]
            ws.append([EventID, RegistrationRequired, InPersonSeats, OnlineSeats, ConfirmedRegistrations, 
                       WaitlistRegistrations, CancelledRegistrations, AnticipatedAttendance, ActualAttendance_InPerson, 
                       ActualAttendance_Online, ConfirmedAttendance])
            
            # Add to Event Program Table
            classify_event_program_by_title(wb, EventID, Title, CategoriesList)

    # If needed, create Matrix Map Settings:
    # This Excel Worksheet allows the Users of the Matrix map to specify how they want to calculate the Impact and Cost
    if not os.path.exists("Matrix Map Settings.xlsx"):
        wb2 = create_matrix_map_settings()

        for category in CategorySet:
            ws = wb2["Category Options"]
            if category:
                ws.append([category])

        for audience in AudienceSet:
            ws = wb2["Audience Options"]
            ws.append([audience])

        programs = ["Story Time","Entertainment","Makerspace and Workshop",
                    "Fitness and Wellness","Tech Support","Language and Culture",
                    "Music and Film","Book Club","Discovery Center",
                    "Life Skills and Community Resource","Nature and Home","Local History & Archives",
                    "Genealogy Services","Important Meeting","Reader's Advisory"]
        
        for program in programs:
            ws = wb2["Program Options"]
            ws.append([program])

        wb2.save("Matrix Map Settings.xlsx")
        wb2.close()

    # close file
    f.close()

from openpyxl import load_workbook

def create_simple_map():
    wb = load_workbook("Matrix Map Dataset.xlsx", data_only=True)

    # Mapping from EventID to ProgramType
    EventID_To_Program = {}
    ws = wb["EventProgram"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        EventID = row[0]
        ProgramType = row[1]  # Assuming second column has ProgramType
        if EventID not in EventID_To_Program:
            EventID_To_Program[EventID] = set()
        EventID_To_Program[EventID].add(ProgramType)

    # Program statistics
    Program_Info = {}
    # {ProgramType: [EventCount, TotalAttendance, ParticipationRate, TotalDuration, TotalStaffTime, AvgDuration, AvgStaffTime]}

    # Process EventParticipation sheet
    ws = wb["EventParticipation"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        EventID = row[0]
        if EventID not in EventID_To_Program:
            print(f"Warning: EventID {EventID} not found in EventID_To_Program mapping.")
            continue

        for Program in EventID_To_Program[EventID]:
            In_Person_Attendance = int(row[8] or 0)
            Online_Attendance = int(row[9] or 0)
            Total_Attendance = In_Person_Attendance + Online_Attendance

            if Program not in Program_Info:
                Program_Info[Program] = [0, 0, 0, 0, 0, 0, 0]

            Program_Info[Program][0] += 1  # EventCount
            Program_Info[Program][1] += Total_Attendance  # TotalAttendance

    # Process EventTimes sheet
    ws = wb["EventTimes"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        EventID = row[0]
        if EventID not in EventID_To_Program:
            print(f"Warning: EventID {EventID} not found in EventID_To_Program mapping.")
            continue

        for Program in EventID_To_Program[EventID]:

            # In EventTimes loop
            Duration = float(row[8] or 0)
            StaffTime = float(row[9] or 0)


            if Program not in Program_Info:
                Program_Info[Program] = [0, 0, 0, 0, 0, 0, 0]

            Program_Info[Program][3] += Duration  # TotalDuration
            Program_Info[Program][4] += StaffTime  # TotalStaffTime

    # Create a Workbook
    new_wb = Workbook()
    # Give the Workbook The Following Worksheets (avg)
    workbook_setup = {  "Program Statistics": ["Program Type", "Event Count", "Total Attendance", "Participation Rate", 
                                               "Total Duration", "Average Duration", "Total Staff Time", "Average Staff Time"],
    }

    # Add Worksheets and Headers
    for sheet_name, headers in workbook_setup.items():
        ws = new_wb.create_sheet(title=sheet_name)
        ws.append(headers)

    # remove the first sheet
    new_wb.remove(new_wb["Sheet"])
    ws = new_wb["Program Statistics"]
    # For Columns A-H, make the column width 20
    for col in range(1, 9):  # 1 to 8 inclusive for A to H
        col_letter = chr(64 + col)  # Convert 1->A, 2->B, ..., 8->H
        ws.column_dimensions[col_letter].width = 20

    for Program, stats in Program_Info.items():
        Total_Events = stats[0]
        Total_Attendance = stats[1]
        Total_Duration = stats[3]
        Total_StaffTime = stats[4]

        ParticipationRate = Total_Attendance / Total_Events if Total_Events else 0
        AvgDuration = Total_Duration / Total_Events if Total_Events else 0
        AvgStaffTime = Total_StaffTime / Total_Events if Total_Events else 0

        stats[2] = ParticipationRate
        stats[5] = AvgDuration
        stats[6] = AvgStaffTime

        new_row = [Program]
        new_row.extend(stats)
        ws.append(new_row)
    
    wb = load_workbook("Matrix Map Settings.xlsx", data_only=False)
    ws1 = wb["Program Options"]
    ws2 = new_wb.create_sheet("Program Options")

    # Copy cell values
    for row in ws1.iter_rows():
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    # Copy column widths
    for col in ws1.column_dimensions:
        ws2.column_dimensions[col].width = ws1.column_dimensions[col].width

    # Copy row heights
    for row in ws1.row_dimensions:
        ws2.row_dimensions[row].height = ws1.row_dimensions[row].height

    new_wb.save("Simple Map.xlsx")
    

def main():
    # Step 1: Search for Data Source File 
    source_file = find_data_source_file()
    print(f"Using data source file: {source_file}")

    # Step 2: Create Dataset Template
    wb = create_dataset_template()
    print(f"Created dataset template: {wb}")

    # Step 3: Check for Existing Dataset and delete it
    if os.path.exists("Matrix Map Dataset.xlsx"):
        os.remove("Matrix Map Dataset.xlsx")
        print("Deleted existing dataset file.")

    # Step 4: Read CSV and Populate Workbook
    read_csv_and_populate_workbook(wb, source_file)
    print("Populated workbook with data from CSV.")

    # Step 5: Save the Workbook
    wb.save("Matrix Map Dataset.xlsx")
    print("Saved workbook as 'Matrix Map Dataset.xlsx'.")

    # Step 6: Close the workbook
    wb.close()

import time
if __name__ == "__main__":
    start_time = time.time()
    main()
    # create_simple_map()
    end_time = time.time()
    print(f"Finished in {end_time - start_time:.3f} seconds.")
    input("Press Enter to Close the Program:")