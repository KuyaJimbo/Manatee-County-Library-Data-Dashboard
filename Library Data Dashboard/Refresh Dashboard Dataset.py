# Import Helper Functions From Python Libraries
from openpyxl import load_workbook, Workbook
import os, re

# =============================================================================
# CONFIGURATION SECTION - Easy to modify mappings and settings
# =============================================================================

MONTHS = ["October", "November", "December", "January", "February", "March", 
          "April", "May", "June", "July", "August", "September"]

EXPECTED_FILES = ['ILL.xlsx', 'Digital Information.xlsx', 'Tech Statistics.xlsx', 'Summary Usage Report.xlsx']

# Cell mappings for different data types
GENERAL_STATISTICS_CELLS = {
    'Total Patrons': 'F8',
    'Volunteers Hours': 'F10',
    'Volunteens Hours': 'F11',
    'Subtotal Hours Worked': 'F12',
    'Public Service Hours': 'F13',
    'Total Hours': 'F14',
    'Paid Non-Library Staff Hours': 'F15',
    'Total Volunteers': 'F17',
    'Reference Count': 'F19',
    'Virtual Reference': 'F20',
    'Staff Receiving': 'F22',
    'Staff Hours': 'F23',
    'Patrons Receiving': 'F24',
    'Hours With Patrons': 'F25',
    'Voter Registration Count': 'F27',
    'In House Count': 'F29',
    'Hours Open Total': 'F31'
}

LITTLE_DISCOVERY_CELLS = {
    'Total Patrons': 'D6'
}

DIGITAL_INFO_CELLS = {
    'Digital Materials': 'E4',
    'Digital Circulation': 'E5',
    'Database Use': 'E6',
    'Current Library Card Holders': 'E7',
    'Residential Card Holders': 'E9',
    'Non Residential Card Holders': 'E10',
    'Avg Hold Time': 'E11',
    'Circulation of Adult Materials': 'E12',
    'Circulation of Youth Materials': 'E13',
    'Other Circulating Materials': 'E14',
    'Physical Item Circulation Total': 'E15',
    'Total Electronic Content Use': 'E16',
    'Total Collection Use': 'E17'
}

TECH_STATS_CELLS = {
    # Row numbers will be added dynamically
    'Check Outs': 'G',  
    'Check Ins': 'I',
    'Total Volumes Available': 'O',
    'New Library Card Holders': 'U'
}

# Computer & Study Room Usage
USAGE_CELLS = {
    # Row numbers will be added dynamically
    'Total Computer Usage': 'M',  
    'Hours Booked': 'T',
    'Total Bookings': 'V',
    'Unique Users': 'X',
    'Number Of Rooms': 'Z'
}

ILL_CELLS = {
    'Borrowed': 'C5:N5',
    'Supplied': 'C6:N6'
}

PROGRAMMING_CATEGORIES = ["In-House", "Outreach", "Virtual", "Self-Directed"]

NON_LIBRARY_USE_CELLS = {
    'Total Groups': 'H34',
    'Total Attendance': 'K34'
}

# Column headers for each worksheet
WORKSHEET_COLUMNS = {
    'General Statistics': ['Location', 'Date', 'Month Name', 'Year', 'Total Patrons', 
                          'Volunteers Hours', 'Volunteens Hours', 'Subtotal Hours Worked', 
                          'Public Service Hours', 'Total Hours', 'Paid Non-Library Staff Hours', 
                          'Total Volunteers', 'Reference Count', 'Virtual Reference', 
                          'Staff Receiving', 'Staff Hours', 'Patrons Receiving', 
                          'Hours With Patrons', 'Voter Registration Count', 'In House Count', 
                          'Hours Open Total'],
    
    'Programming': ['Location', 'Date', 'Month Name', 'Year', 'Category', 'Age Group', 
                   'Total Groups/Sessions', 'Total Attendance'],
    
    'Digital Information': ['Date', 'Month Name', 'Year', 'Digital Materials', 
                           'Digital Circulation', 'Database Use', 'Current Library Card Holders',
                           'Residential Card Holders', 'Non Residential Card Holders', 
                           'Avg Hold Time', 'Circulation of Adult Materials',
                           'Circulation of Youth Materials', 'Other Circulating Materials',
                           'Physical Item Circulation Total', 'Total Electronic Content Use',
                           'Total Collection Use'],
    
    'ILL': ['Date', 'Month Name', 'Year', 'Borrowed', 'Supplied'],
    
    'Tech Statistics': ['Location', 'Date', 'Month Name', 'Year', 'Check Outs', 'Check Ins', 
                       'Total Volumes Available', 'New Library Card Holders'],
    
    'Tech Statistics pt2': ['Date', 'Month Name', 'Year', 'Reserve Taken', 'Reserve Filled', 
                           'Total Titles Available', 'Patrons Logins', 'Total PAC Searches'],
    
    'Computer & Study Room Usage': ['Location', 'Date', 'Month Name', 'Year', 'Total Computer Usage',
                         'Hours Booked', 'Total Bookings', 
                         'Unique Users', 'Number Of Rooms'],
    
    'Branch Legend': ['Name', 'Location']
}

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def detect_branch_files(folder_path):
    """Detect all Excel files ending with 'Branch.xlsx' and create branch legend data."""
    branch_files = []
    branch_legend_data = []
    
    for filename in os.listdir(folder_path):
        if filename.endswith(' Branch.xlsx'):
            branch_files.append(filename)
            # Remove ' Branch.xlsx' to get the location name
            location_name = filename.replace(' Branch.xlsx', '')
            branch_legend_data.append([filename.replace('.xlsx', ''), location_name])
    
    return branch_files, branch_legend_data

def get_library_locations_from_files(branch_files):
    """Extract library location names from branch files for tech statistics mapping."""
    locations = []
    for filename in branch_files:
        location = filename.replace(' Branch.xlsx', '')
        locations.append(location)
    return locations

def get_month_number_from_name(month_name):
    """Convert month name to month number as 2-character string."""
    month_number = (MONTHS.index(month_name) + 9) % 12 + 1
    return f"{month_number:02d}"

def get_date_string(year, month_name):
    """Create date string in format YYYY-MM-01 00:00:00."""
    month_number = get_month_number_from_name(month_name)
    return f"{year}-{month_number}-01 00:00:00"

def get_year_from_month(month_name, start_year):
    """Determine the correct year based on the month and fiscal year start."""
    if month_name in ["January", "February", "March", "April", "May", "June", "July", "August", "September"]:
        return start_year + 1
    return start_year

def clean_data_row(row, skip_columns=4):
    """Clean data row by replacing None values with 0 and checking for string values in numeric columns."""
    clean_list = [0 if x is None else x for x in row]
    # Check if numeric columns (after specified number of columns) contain strings
    if any(isinstance(item, str) for item in clean_list[skip_columns:]):
        raise TypeError("Some data is not numerical.")
    return clean_list

def safe_append_library_row(worksheet, row, location, month_name, year, data_type, skip_columns=4):
    """Safely append a row to worksheet with error handling."""
    try:
        clean_row = clean_data_row(row, skip_columns)
        worksheet.append(clean_row)
    except Exception as e:
        print(f"{data_type} in {location} {month_name} {year}: {e}")

def safe_append_row(worksheet, row, month_name, year, data_type, skip_columns=4):
    """Safely append a row to worksheet with error handling."""
    try:
        clean_row = clean_data_row(row, skip_columns)
        worksheet.append(clean_row)
    except Exception as e:
        print(f"{data_type} in {month_name} {year}: {e}")

# =============================================================================
# WORKSHEET CREATION AND SETUP
# =============================================================================

def create_master_dataset():
    """Create the master workbook with all required worksheets and headers."""
    new_wb = Workbook()
    
    # Create worksheets
    worksheets = {}
    ws_names = list(WORKSHEET_COLUMNS.keys())
    
    # First worksheet (active)
    ws1 = new_wb.active
    ws1.title = ws_names[0]
    worksheets[ws_names[0]] = ws1
    
    # Create remaining worksheets
    for ws_name in ws_names[1:]:
        worksheets[ws_name] = new_wb.create_sheet(ws_name)
    
    # Add column headers to each worksheet
    for ws_name, columns in WORKSHEET_COLUMNS.items():
        worksheets[ws_name].append(columns)
    
    return new_wb, worksheets

def populate_legend_worksheets(worksheets, branch_legend_data):
    """Populate the Branch Legend and Age Group Legend worksheets."""
    # Populate Branch Legend
    for branch_data in branch_legend_data:
        worksheets['Branch Legend'].append(branch_data)

# =============================================================================
# DATA EXTRACTION FUNCTIONS
# =============================================================================

def extract_general_statistics(sheet, location, month_name, year, date):
    """Extract general statistics data from a library worksheet."""
    if location == "Little Discovery Center":  # Little Discovery Center special case
        total_patrons = sheet[LITTLE_DISCOVERY_CELLS['Total Patrons']].value
        return [location, date, month_name, year, total_patrons] + [0] * 16
    else:
        # Extract all general statistics using the cell mapping
        row_data = [location, date, month_name, year]
        for metric, cell in GENERAL_STATISTICS_CELLS.items():
            row_data.append(sheet[cell].value)
        return row_data
 
def extract_programming_data(sheet, location, month_name, year, date):
    """Extract programming data from a library worksheet."""
    programming_rows = []
    
    # Regular programming categories
    for index, category in enumerate(PROGRAMMING_CATEGORIES):
        row_number = 39 + (index * 9)
        columns = ['H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
        
        for col_index in range(0, len(columns), 2):
            age_group = sheet[columns[col_index] + str(row_number)].value
            total_groups = sheet[columns[col_index] + str(row_number + 2)].value
            total_attendance = sheet[columns[col_index + 1] + str(row_number + 2)].value
            
            row = [location, date, month_name, year, category, age_group, total_groups, total_attendance]
            programming_rows.append(row)
    
    # Non Library Use of Facilities
    category = 'Non Library Use of Facilities'
    age_group = 'N/A'
    total_groups = sheet[NON_LIBRARY_USE_CELLS['Total Groups']].value
    total_attendance = sheet[NON_LIBRARY_USE_CELLS['Total Attendance']].value
    
    row = [location, date, month_name, year, category, age_group, total_groups, total_attendance]
    programming_rows.append(row)
    
    return programming_rows

def extract_digital_info(sheet, month_name, year, date):
    """Extract digital information data from worksheet."""
    row_data = [date, month_name, year]
    for metric, cell in DIGITAL_INFO_CELLS.items():
        row_data.append(sheet[cell].value)
    return row_data

def extract_ill_data(sheet, start_year):
    """Extract Inter Library Loan data from worksheet."""
    borrowed_list = [cell.value for row in sheet[ILL_CELLS['Borrowed']] for cell in row]
    supplied_list = [cell.value for row in sheet[ILL_CELLS['Supplied']] for cell in row]
    
    ill_rows = []
    for index in range(12):
        year = start_year if index < 3 else start_year + 1
        month_number = (index + 9) % 12 + 1
        month_name = MONTHS[index]  # This directly uses the index to get the correct month
        date = f"{year}-{month_number:02d}-01 00:00:00"
        
        borrowed = borrowed_list[index]
        supplied = supplied_list[index]
        
        row = [date, month_name, str(year), borrowed, supplied]
        ill_rows.append(row)
    
    return ill_rows

def extract_tech_statistics(sheet, month_name, year, date):
    """Extract technology statistics from worksheet."""
    tech_rows = []
    
    # Process libraries based on detected branch files
    index = 5
    while sheet[f"B{index}"].value != 'Total':
        location = sheet[f"B{index}"].value
        check_outs = sheet[f"{TECH_STATS_CELLS['Check Outs']}{index}"].value
        check_ins = sheet[f"{TECH_STATS_CELLS['Check Ins']}{index}"].value
        total_volumes = sheet[f"{TECH_STATS_CELLS['Total Volumes Available']}{index}"].value
        new_cards = sheet[f"{TECH_STATS_CELLS['New Library Card Holders']}{index}"].value
        
        row = [location, date, month_name, year, check_outs, check_ins, total_volumes, new_cards]
        tech_rows.append(row)
        index += 1
    
    return tech_rows

def extract_computer_study_room_usage(sheet, month_name, year, date):
    """Extract technology statistics from worksheet."""
    usage_rows = []
    
    # Process libraries based on detected branch files
    index = 7
    while sheet[f"J{index}"].value != 'Total':
        location = sheet[f"J{index}"].value
        total_computer_usage = sheet[f"{USAGE_CELLS['Total Computer Usage']}{index}"].value
        hours_booked = sheet[f"{USAGE_CELLS['Hours Booked']}{index}"].value
        total_bookings = sheet[f"{USAGE_CELLS['Total Bookings']}{index}"].value
        unique_users = sheet[f"{USAGE_CELLS['Unique Users']}{index}"].value
        num_of_rooms = sheet[f"{USAGE_CELLS['Number Of Rooms']}{index}"].value

        row = [location, date, month_name, year, total_computer_usage, hours_booked, total_bookings, unique_users, num_of_rooms]
        usage_rows.append(row)
        index += 1
        if index == 30:
            break

    return usage_rows

def extract_tech_statistics_pt2(sheet, month_name, year, date):
    """Extract additional tech statistics from worksheet."""
    tech_stats_2_data = [cell.value for row in sheet["AA4:AA8"] for cell in row]
    return [date, month_name, year] + tech_stats_2_data

# =============================================================================
# MAIN PROCESSING FUNCTIONS
# =============================================================================

def process_library_file(wb, worksheets, filename, start_year):
    """Process a single library Excel file."""
    for sheet_name in wb.sheetnames:
        if sheet_name not in MONTHS:
            print(f"Ignored: {sheet_name} sheet from {filename}")
            continue
        
        sheet = wb[sheet_name]
        location = sheet["B4"].value
        
        if location == "Month/Year":
            location = "Little Discovery Center"
        year = get_year_from_month(sheet_name, start_year)
        date = get_date_string(year, sheet_name)
        
        # Extract and append general statistics
        general_stats = extract_general_statistics(sheet, location, sheet_name, year, date)
        safe_append_library_row(worksheets['General Statistics'], general_stats, 
                       location, sheet_name, year, "General Statistics")

        # Extract and append programming data (skip for Little Discovery Center)
        if location != "Month/Year":
            programming_data = extract_programming_data(sheet, location, sheet_name, year, date)
            for prog_row in programming_data:
                safe_append_library_row(worksheets['Programming'], prog_row, 
                               location, sheet_name, year, "Programming", skip_columns=6)

def process_digital_info_file(wb, worksheets, start_year):
    """Process Digital Information Excel file."""
    for sheet_name in wb.sheetnames:
        if sheet_name not in MONTHS:
            continue
        
        sheet = wb[sheet_name]
        year = get_year_from_month(sheet_name, start_year)
        date = get_date_string(year, sheet_name)
        
        digital_data = extract_digital_info(sheet, sheet_name, year, date)
        safe_append_row(worksheets['Digital Information'], digital_data, sheet_name, year, "Digital Information")


def process_tech_stats_file(wb, worksheets, start_year):
    """Process Tech Statistics Excel file."""
    for sheet_name in wb.sheetnames:
        if sheet_name not in MONTHS:
            continue
        
        sheet = wb[sheet_name]
        year = get_year_from_month(sheet_name, start_year)
        date = get_date_string(year, sheet_name)
        
        # Process main tech statistics
        tech_data = extract_tech_statistics(sheet, sheet_name, year, date)
        for tech_row in tech_data:
            safe_append_library_row(worksheets['Tech Statistics'], tech_row, 
                           tech_row[0], sheet_name, year, "Tech Statistics")
        
        # Process tech statistics part 2
        tech_stats_2 = extract_tech_statistics_pt2(sheet, sheet_name, year, date)
        safe_append_row(worksheets['Tech Statistics pt2'], tech_stats_2, sheet_name, year, "Tech Statistics pt2")

def process_library_usage(wb, worksheets, start_year):
    """Process computer and study room usage data."""
    for sheet_name in wb.sheetnames:
        if sheet_name not in MONTHS:
            continue

        sheet = wb[sheet_name]
        year = get_year_from_month(sheet_name, start_year)
        date = get_date_string(year, sheet_name)

        computer_study_room_usage = extract_computer_study_room_usage(sheet, sheet_name, year, date)
        for usage_row in computer_study_room_usage:
            safe_append_library_row(worksheets['Computer & Study Room Usage'], usage_row, usage_row[0], sheet_name, year, "Computer & Study Room Usage")

def delete_master_dataset():
    """Remove existing master dataset file if it exists."""
    if os.path.exists('MasterDataset.xlsx'):
        os.remove('MasterDataset.xlsx')
        print('Removed existing Master Dataset')

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main execution function."""
    print("Starting Master Dataset Creation...")
    
    # Remove existing file and create new workbook
    delete_master_dataset()
    new_wb, worksheets = create_master_dataset()

    # MODIFIED: Get parent directory (go up one level from current script location)
    script_directory = os.getcwd()  # This is now the Dashboard folder
    parent_directory = os.path.dirname(script_directory)  # Go up one level

    folders = [folder for folder in os.listdir(parent_directory) 
               if (os.path.isdir(os.path.join(parent_directory, folder)) and 
                   re.match(r"^October 2\d{3} - September 2\d{3}$", folder))]
    
    print(f"Found {len(folders)} fiscal year folders to process.")
    
    # IMPROVED: Collect all unique branches across all fiscal years
    all_branch_legend_data = {}  # Use dict to avoid duplicates
    
    for folder_name in folders:
        print(f"\nProcessing folder: {folder_name}")
        
        folder_path = os.path.join(parent_directory, folder_name)
        
        # Detect branch files dynamically
        branch_files, branch_legend_data = detect_branch_files(folder_path)
        library_locations = get_library_locations_from_files(branch_files)
        
        # IMPROVED: Collect branch legend data from ALL folders
        for branch_data in branch_legend_data:
            branch_name, location_name = branch_data
            all_branch_legend_data[branch_name] = location_name
        
        # Check for missing files
        all_expected_files = EXPECTED_FILES + branch_files
        missing_files = [f"{file} not found in {folder_name}" 
                        for file in all_expected_files 
                        if file not in os.listdir(folder_path)]

        if missing_files:
            print('\n'.join(missing_files))
        
        start_year = int(folder_name.split(" - ")[0].split()[1])
        
        # Process each file in the folder
        for filename in os.listdir(folder_path):
            
            # Skip temporary Excel files (start with ~$)
            if filename.startswith('~$'):
                print(f"Ignored: {filename} is currently in use, refresh again later to see new updates with current data.")
                continue
            
            full_path = os.path.join(folder_path, filename)
            
            # Process branch files dynamically
            if filename in branch_files:
                wb = load_workbook(full_path, data_only=True)
                process_library_file(wb, worksheets, filename, start_year)
                print(f"Processed: {filename}")
                
            elif filename == 'ILL.xlsx':
                wb = load_workbook(full_path, data_only=True)
                sheet = wb.active
                ill_data = extract_ill_data(sheet, start_year)
                for ill_row in ill_data:
                    safe_append_library_row(worksheets['ILL'], ill_row, 
                                   "System-wide", ill_row[1], ill_row[2], "ILL", skip_columns=3)
                print(f"Processed: {filename}")
                
            elif filename == 'Digital Information.xlsx':
                wb = load_workbook(full_path, data_only=True)
                process_digital_info_file(wb, worksheets, start_year)
                print(f"Processed: {filename}")
                
            elif filename == 'Tech Statistics.xlsx':
                wb = load_workbook(full_path, data_only=True)
                process_tech_stats_file(wb, worksheets, start_year)
                print(f"Processed: {filename}")
                
            elif filename == 'Summary Usage Report.xlsx':
                wb = load_workbook(full_path, data_only=True)
                process_library_usage(wb, worksheets, start_year)
                print(f"Processed: {filename}")
                
            else:
                if filename.endswith('.xlsx'):
                    print(f"Ignored: {filename}")
    
    # IMPROVED: Convert dict back to list format for populate_legend_worksheets
    final_branch_legend_data = [[branch_name, location_name] 
                               for branch_name, location_name in all_branch_legend_data.items()]

    # Populate legend worksheets
    populate_legend_worksheets(worksheets, final_branch_legend_data)
    
    # Save the final dataset
    new_wb.save('MasterDataset.xlsx')
    print(f"\nCompleted! Master Dataset created with {len(worksheets)} worksheets.")
    input("Press Enter to exit...")

if __name__ == "__main__":
    main()